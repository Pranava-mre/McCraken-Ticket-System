import io
import os
import csv
import json
import re
import hmac
import time
import threading
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from datetime import timedelta
from urllib.parse import parse_qs, urlparse, unquote
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import base64
import re
import psycopg2
import pyodbc
import requests
from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase.pdfmetrics import stringWidth
from psycopg2 import IntegrityError
from psycopg2.extras import RealDictCursor

BlobServiceClient = None
ContentSettings = None

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
DATA_DIR = BASE_DIR / "data"
SCHEMA_PATH = BASE_DIR / "schema.sql"


def resolve_storage_dir(env_var_name, default_relative_path):
    configured_path = os.getenv(env_var_name, "").strip()
    if configured_path:
        resolved = Path(configured_path).expanduser()
        if not resolved.is_absolute():
            resolved = BASE_DIR / resolved
    else:
        resolved = BASE_DIR / default_relative_path
    resolved.mkdir(parents=True, exist_ok=True)
    return resolved


PDF_DIR = resolve_storage_dir("TICKETS_PDF_DIR", "tickets_pdf")
REPORT_PDF_DIR = resolve_storage_dir("REPORT_PDF_DIR", "reports_pdf")


app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "local-dev-secret-key")
app.permanent_session_lifetime = timedelta(hours=8)

PROCESS_TZ = os.getenv("TZ", "").strip()
if PROCESS_TZ and hasattr(time, "tzset"):
    os.environ["TZ"] = PROCESS_TZ
    try:
        time.tzset()
    except Exception:
        app.logger.warning("Could not apply process timezone '%s'.", PROCESS_TZ)

APP_TIMEZONE = os.getenv("APP_TIMEZONE", PROCESS_TZ or "UTC").strip() or "UTC"
try:
    APP_TZ = ZoneInfo(APP_TIMEZONE)
except ZoneInfoNotFoundError:
    APP_TZ = ZoneInfo("UTC")
    app.logger.warning("Invalid APP_TIMEZONE '%s'. Falling back to UTC.", APP_TIMEZONE)


def app_now():
    return datetime.now(APP_TZ)

APP_USERNAME = os.getenv("APP_USERNAME", "").strip()
APP_PASSWORD = os.getenv("APP_PASSWORD", "").strip()

_db_init_lock = threading.Lock()
_db_initialized = False

AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "").strip()
AZURE_STORAGE_CONTAINER = os.getenv("AZURE_STORAGE_CONTAINER", "ticket-pdfs").strip() or "ticket-pdfs"
AZURE_TICKETS_BLOB_PREFIX = os.getenv("AZURE_TICKETS_BLOB_PREFIX", "tickets").strip().strip("/")
AZURE_REPORTS_BLOB_PREFIX = os.getenv("AZURE_REPORTS_BLOB_PREFIX", "reports").strip().strip("/")
AZURE_DOWNLOADS_BLOB_PREFIX = os.getenv("AZURE_DOWNLOADS_BLOB_PREFIX", "downloads").strip().strip("/")
AZURE_TICKET_IMAGES_BLOB_PREFIX = os.getenv("AZURE_TICKET_IMAGES_BLOB_PREFIX", "ticket-images").strip().strip("/")
AZURE_JOBS_CACHE_BLOB_NAME = os.getenv("AZURE_JOBS_CACHE_BLOB_NAME", "jobs_cache/jobs_cache.csv").strip().strip("/")
CREDIT_CARD_CUSTOMER_MATCH = os.getenv("CREDIT_CARD_CUSTOMER_MATCH", "credit card").strip() or "credit card"
CREDIT_CARD_REPORT_API_KEY = os.getenv("CREDIT_CARD_REPORT_API_KEY", "").strip()
CREDIT_CARD_REPORT_SAS_MINUTES = int(os.getenv("CREDIT_CARD_REPORT_SAS_MINUTES", "180").strip() or "180")
AZURE_IMAGE_SAS_MINUTES = int(os.getenv("AZURE_IMAGE_SAS_MINUTES", "180").strip() or "180")
RFID_EVENT_API_KEY = os.getenv("RFID_EVENT_API_KEY", "").strip()
RFID_WEBHOOK_USERNAME = os.getenv("RFID_WEBHOOK_USERNAME", "").strip()
RFID_WEBHOOK_PASSWORD = os.getenv("RFID_WEBHOOK_PASSWORD", "").strip()
NOTIFICATIONS_ENABLED = os.getenv("NOTIFICATIONS_ENABLED", "0").strip().lower() in {"1", "true", "yes", "on"}
app.config["NOTIFICATIONS_ENABLED"] = NOTIFICATIONS_ENABLED
try:
    RFID_NOTIFICATION_COOLDOWN_SECONDS = max(
        0,
        int(os.getenv("RFID_NOTIFICATION_COOLDOWN_SECONDS", "60").strip() or "60"),
    )
except ValueError:
    RFID_NOTIFICATION_COOLDOWN_SECONDS = 60


def env_flag(name, default=False):
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


AUTO_DB_BOOTSTRAP = env_flag("AUTO_DB_BOOTSTRAP", True)
PG_CONNECT_TIMEOUT = int(os.getenv("PG_CONNECT_TIMEOUT", "8").strip() or "8")
MAX_JOB_OPTIONS = int(os.getenv("MAX_JOB_OPTIONS", "2000").strip() or "2000")

WAVE_IMAGE_CAPTURE_ENABLED = env_flag("WAVE_IMAGE_CAPTURE_ENABLED", True)
WAVE_CLOUD_BASE_URL = os.getenv("WAVE_CLOUD_BASE_URL", "https://sync.wavevms.com").strip().rstrip("/")
WAVE_SYSTEM_ID = os.getenv("WAVE_SYSTEM_ID", "").strip()
WAVE_USERNAME = os.getenv("WAVE_USERNAME", "").strip()
WAVE_PASSWORD = os.getenv("WAVE_PASSWORD", "").strip()
WAVE_CAMERA_ID = os.getenv("WAVE_CAMERA_ID", "").strip()
WAVE_SERVER_GUID = os.getenv("WAVE_SERVER_GUID", "").strip()
WAVE_IMAGE_RESOLUTION = os.getenv("WAVE_IMAGE_RESOLUTION", "1024x452").strip() or "1024x452"
WAVE_FOOTAGE_DURATION_SECONDS = max(1, int(os.getenv("WAVE_FOOTAGE_DURATION_SECONDS", "8").strip() or "8"))

_wave_session_lock = threading.Lock()
_wave_session_cache = {
    "access_token": "",
    "relay_root": "",
    "expires_at_unix": 0,
}


def _wave_is_configured():
    required = [WAVE_CLOUD_BASE_URL, WAVE_SYSTEM_ID, WAVE_USERNAME, WAVE_PASSWORD, WAVE_CAMERA_ID]
    return all(bool(v) for v in required)


def _wave_ticket_time_to_ms(ticket_time_text):
    text = str(ticket_time_text or "").strip()
    if not text:
        raise ValueError("Empty ticket timestamp.")

    if re.fullmatch(r"[0-9]{13}", text):
        return int(text)
    if re.fullmatch(r"[0-9]{10}(?:\.[0-9]+)?", text):
        return int(float(text) * 1000)

    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=APP_TZ)
    return int(dt.timestamp() * 1000)


def _wave_fetch_access_token():
    response = requests.post(
        f"{WAVE_CLOUD_BASE_URL}/cdb/oauth2/token",
        json={
            "grant_type": "password",
            "response_type": "token",
            "client_id": "3rdParty",
            "scope": f"cloudSystemId={WAVE_SYSTEM_ID}",
            "username": WAVE_USERNAME,
            "password": WAVE_PASSWORD,
        },
        timeout=30,
    )
    if response.status_code != 200:
        raise RuntimeError(f"WAVE auth failed ({response.status_code}): {response.text[:300]}")

    payload = response.json() if response.content else {}
    token = str(payload.get("access_token") or "").strip()
    if not token:
        raise RuntimeError("WAVE auth response missing access_token.")
    if not token.startswith("nxcdb-"):
        token = "nxcdb-" + token

    expires_in = int(payload.get("expires_in") or 3600)
    expires_at_unix = int(time.time()) + max(60, expires_in - 60)
    return token, expires_at_unix


def _wave_fetch_relay_root(access_token):
    base_url = f"https://{WAVE_SYSTEM_ID}.relay.vmsproxy.com"
    response = requests.get(
        f"{base_url}/rest/v4/login/sessions/{access_token}",
        timeout=30,
        allow_redirects=True,
    )
    if response.status_code != 200:
        raise RuntimeError(f"WAVE relay lookup failed ({response.status_code}).")

    marker = "/rest/v4/login/sessions/"
    url = str(response.url or "")
    idx = url.find(marker)
    if idx < 0:
        raise RuntimeError("WAVE relay URL format unexpected.")
    return url[:idx]


def _wave_get_or_refresh_session(force_refresh=False):
    with _wave_session_lock:
        now_unix = int(time.time())
        has_valid_cached = (
            bool(_wave_session_cache.get("access_token"))
            and bool(_wave_session_cache.get("relay_root"))
            and int(_wave_session_cache.get("expires_at_unix") or 0) > now_unix
        )
        if has_valid_cached and not force_refresh:
            return _wave_session_cache["access_token"], _wave_session_cache["relay_root"]

        access_token, expires_at_unix = _wave_fetch_access_token()
        relay_root = _wave_fetch_relay_root(access_token)
        _wave_session_cache.update(
            {
                "access_token": access_token,
                "relay_root": relay_root,
                "expires_at_unix": expires_at_unix,
            }
        )
        return access_token, relay_root


def _wave_extract_first_jpeg(stream_response):
    buffer = b""
    for chunk in stream_response.iter_content(chunk_size=4096):
        if not chunk:
            continue
        buffer += chunk
        start = buffer.find(b"\xff\xd8")
        end = buffer.find(b"\xff\xd9")
        if start != -1 and end != -1 and end > start:
            return buffer[start:end + 2]
    return None


def _wave_capture_ticket_image_bytes(created_at_text):
    if not _wave_is_configured():
        return None, "WAVE_NOT_CONFIGURED", "WAVE settings are missing on server."

    try:
        position_ms = _wave_ticket_time_to_ms(created_at_text)
    except Exception as exc:
        return None, "INVALID_TIMESTAMP", f"Invalid ticket timestamp: {exc}"

    for attempt in range(2):
        force_refresh = attempt > 0
        try:
            access_token, relay_root = _wave_get_or_refresh_session(force_refresh=force_refresh)
        except Exception as exc:
            return None, "WAVE_AUTH_FAILED", str(exc)

        headers = {"Authorization": f"Bearer {access_token}"}
        playback_response = requests.post(
            f"{relay_root}/rest/v3/login/tickets",
            headers=headers,
            json={},
            timeout=30,
        )
        if playback_response.status_code == 401 and attempt == 0:
            continue
        if playback_response.status_code != 200:
            return None, "PLAYBACK_TICKET_FAILED", f"Playback ticket failed ({playback_response.status_code})."

        playback_token = str((playback_response.json() or {}).get("token") or "").strip()
        if not playback_token:
            return None, "PLAYBACK_TICKET_FAILED", "Playback ticket token missing."

        footage_response = requests.get(
            f"{relay_root}/rest/v4/devices/{WAVE_CAMERA_ID}/footage",
            headers=headers,
            params={
                "startTimeMs": position_ms,
                "endTimeMs": position_ms + (WAVE_FOOTAGE_DURATION_SECONDS * 1000),
                "periodType": "recording",
                "preciseBounds": "true",
                "_ticket": playback_token,
            },
            timeout=30,
        )
        if footage_response.status_code == 401 and attempt == 0:
            continue
        if footage_response.status_code != 200:
            return None, "FOOTAGE_CHECK_FAILED", f"Footage check failed ({footage_response.status_code})."

        footage_json = footage_response.json() if footage_response.content else []
        if not footage_json:
            return None, "NO_FOOTAGE", "No recorded footage at this ticket time."

        # Playback tickets can be single-use in some WAVE setups.
        # Create a new ticket for the media fetch after footage probing.
        media_ticket_response = requests.post(
            f"{relay_root}/rest/v3/login/tickets",
            headers=headers,
            json={},
            timeout=30,
        )
        if media_ticket_response.status_code == 401 and attempt == 0:
            continue
        if media_ticket_response.status_code != 200:
            return None, "PLAYBACK_TICKET_FAILED", f"Media playback ticket failed ({media_ticket_response.status_code})."

        media_playback_token = str((media_ticket_response.json() or {}).get("token") or "").strip()
        if not media_playback_token:
            return None, "PLAYBACK_TICKET_FAILED", "Media playback ticket token missing."

        media_params = {
            "positionMs": position_ms,
            "stream": "primary",
            "resolution": WAVE_IMAGE_RESOLUTION,
            "accurateSeek": "true",
            "_ticket": media_playback_token,
        }
        if WAVE_SERVER_GUID:
            media_params["Server-Guid"] = WAVE_SERVER_GUID

        media_response = requests.get(
            f"{relay_root}/rest/v3/devices/{WAVE_CAMERA_ID}/media.mpjpeg",
            headers=headers,
            params=media_params,
            stream=True,
            timeout=120,
        )
        if media_response.status_code == 401 and attempt == 0:
            continue
        if media_response.status_code != 200:
            body_preview = ""
            try:
                body_preview = (media_response.text or "")[:200]
            except Exception:
                body_preview = ""
            msg = f"Image capture failed ({media_response.status_code})."
            if body_preview:
                msg = f"{msg} {body_preview}"
            return None, "IMAGE_CAPTURE_FAILED", msg

        jpeg_bytes = _wave_extract_first_jpeg(media_response)
        media_response.close()
        if not jpeg_bytes:
            return None, "IMAGE_CAPTURE_FAILED", "No JPEG frame found in stream."
        return jpeg_bytes, "", ""

    return None, "WAVE_AUTH_FAILED", "WAVE session refresh did not resolve authentication." 


def _wave_capture_live_image_bytes():
    if not _wave_is_configured():
        return None, "WAVE_NOT_CONFIGURED", "WAVE settings are missing on server."

    for attempt in range(2):
        force_refresh = attempt > 0
        try:
            access_token, relay_root = _wave_get_or_refresh_session(force_refresh=force_refresh)
        except Exception as exc:
            return None, "WAVE_AUTH_FAILED", str(exc)

        headers = {"Authorization": f"Bearer {access_token}"}
        playback_response = requests.post(
            f"{relay_root}/rest/v3/login/tickets",
            headers=headers,
            json={},
            timeout=30,
        )
        if playback_response.status_code == 401 and attempt == 0:
            continue
        if playback_response.status_code != 200:
            return None, "PLAYBACK_TICKET_FAILED", f"Playback ticket failed ({playback_response.status_code})."

        playback_token = str((playback_response.json() or {}).get("token") or "").strip()
        if not playback_token:
            return None, "PLAYBACK_TICKET_FAILED", "Playback ticket token missing."

        media_params = {
            "stream": "primary",
            "resolution": WAVE_IMAGE_RESOLUTION,
            "_ticket": playback_token,
        }
        if WAVE_SERVER_GUID:
            media_params["Server-Guid"] = WAVE_SERVER_GUID

        media_response = requests.get(
            f"{relay_root}/rest/v4/devices/{WAVE_CAMERA_ID}/media.mpjpeg",
            headers=headers,
            params=media_params,
            stream=True,
            timeout=120,
        )
        if media_response.status_code == 401 and attempt == 0:
            continue
        if media_response.status_code != 200:
            body_preview = ""
            try:
                body_preview = (media_response.text or "")[:200]
            except Exception:
                body_preview = ""
            msg = f"Live snapshot failed ({media_response.status_code})."
            if body_preview:
                msg = f"{msg} {body_preview}"
            return None, "IMAGE_CAPTURE_FAILED", msg

        jpeg_bytes = _wave_extract_first_jpeg(media_response)
        media_response.close()
        if not jpeg_bytes:
            return None, "IMAGE_CAPTURE_FAILED", "No JPEG frame found in live stream."
        return jpeg_bytes, "", ""

    return None, "WAVE_AUTH_FAILED", "WAVE session refresh did not resolve authentication."


def upload_ticket_image_to_blob(blob_name, image_bytes):
    blob_service = get_blob_service_client()
    if blob_service is None:
        return None

    container_client = blob_service.get_container_client(AZURE_STORAGE_CONTAINER)
    try:
        container_client.create_container()
    except Exception:
        pass

    blob_client = container_client.get_blob_client(blob_name)
    content_settings = ContentSettings(content_type="image/jpeg") if ContentSettings else None
    blob_client.upload_blob(image_bytes, overwrite=True, content_settings=content_settings)
    return blob_client.url


def _safe_blob_segment(value, default="unknown"):
    raw = str(value or "").strip()
    if not raw:
        raw = default
    safe = re.sub(r"[^a-zA-Z0-9._\-]", "_", raw)
    return safe[:120] or default


def _build_ticket_image_blob_name(ticket_number, created_at_text, material_name):
    year = app_now().year
    material_segment = _safe_blob_segment(material_name, default="material")
    ticket_segment = _safe_blob_segment(ticket_number, default="ticket")
    stamp = _safe_blob_segment(str(created_at_text or ""), default=app_now().strftime("%Y%m%dT%H%M%S"))
    file_name = f"{ticket_segment}_{stamp}.jpg"
    prefix = AZURE_TICKET_IMAGES_BLOB_PREFIX
    return f"{prefix}/{year}/{material_segment}/{file_name}" if prefix else f"{year}/{material_segment}/{file_name}"


def _set_ticket_image_state(db, ticket_id, image_status, image_url="", image_error=""):
    generated_at = app_now().isoformat(timespec="seconds") if image_status == "READY" else None
    with db.cursor() as cursor:
        cursor.execute(
            """
            UPDATE tickets
            SET image_status = %s,
                image_url = %s,
                image_error = %s,
                image_generated_at = %s
            WHERE id = %s
            """,
            (image_status, image_url, image_error, generated_at, ticket_id),
        )


def generate_ticket_image_for_row(db, ticket_row, force=False, capture_mode="historical"):
    ticket_id = int(ticket_row.get("id") or 0)
    if ticket_id <= 0:
        return {"ok": False, "code": "INVALID_TICKET", "message": "Ticket id is missing."}

    existing_url = str(ticket_row.get("image_url") or "").strip()
    if existing_url and not force:
        return {
            "ok": True,
            "already_exists": True,
            "image_url": existing_url,
            "code": "READY",
            "message": "Image already exists.",
        }

    if not WAVE_IMAGE_CAPTURE_ENABLED:
        _set_ticket_image_state(db, ticket_id, image_status="DISABLED", image_url=existing_url, image_error="Image capture is disabled.")
        return {"ok": False, "code": "DISABLED", "message": "Image capture is disabled."}

    mode = str(capture_mode or "historical").strip().lower()
    if mode == "live":
        image_bytes, error_code, error_message = _wave_capture_live_image_bytes()
    else:
        image_bytes, error_code, error_message = _wave_capture_ticket_image_bytes(ticket_row.get("created_at"))
    if not image_bytes:
        status = "NO_FOOTAGE" if error_code == "NO_FOOTAGE" else "ERROR"
        _set_ticket_image_state(db, ticket_id, image_status=status, image_url="", image_error=error_message)
        return {"ok": False, "code": error_code or "ERROR", "message": error_message or "Could not generate image."}

    blob_name = _build_ticket_image_blob_name(
        ticket_number=ticket_row.get("ticket_number"),
        created_at_text=ticket_row.get("created_at"),
        material_name=ticket_row.get("material_name_snapshot"),
    )
    try:
        blob_url = upload_ticket_image_to_blob(blob_name, image_bytes)
    except Exception as exc:
        _set_ticket_image_state(db, ticket_id, image_status="ERROR", image_url="", image_error=f"Blob upload failed: {exc}")
        return {"ok": False, "code": "UPLOAD_FAILED", "message": f"Blob upload failed: {exc}"}

    if not blob_url:
        _set_ticket_image_state(db, ticket_id, image_status="ERROR", image_url="", image_error="Blob upload failed.")
        return {"ok": False, "code": "UPLOAD_FAILED", "message": "Blob upload failed."}

    _set_ticket_image_state(db, ticket_id, image_status="READY", image_url=blob_url, image_error="")
    return {
        "ok": True,
        "already_exists": False,
        "image_url": blob_url,
        "code": "READY",
        "message": "Image generated.",
    }


def get_ticket_image_source_row(db, ticket_id):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                id,
                ticket_number,
                created_at,
                material_name_snapshot,
                image_url,
                image_status,
                image_error
            FROM tickets
            WHERE id = %s
            LIMIT 1
            """,
            (ticket_id,),
        )
        return cursor.fetchone()


def get_blob_service_client():
    global BlobServiceClient, ContentSettings

    if not AZURE_STORAGE_CONNECTION_STRING:
        return None

    if BlobServiceClient is None:
        try:
            from azure.storage.blob import BlobServiceClient as _BlobServiceClient
            from azure.storage.blob import ContentSettings as _ContentSettings

            BlobServiceClient = _BlobServiceClient
            ContentSettings = _ContentSettings
        except ImportError:
            app.logger.warning(
                "Azure Blob connection string is set, but azure-storage-blob package is not installed."
            )
            return None

    try:
        return BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
    except Exception as exc:
        app.logger.warning("Could not initialize Azure Blob client: %s", exc)
        return None


def upload_pdf_to_blob(blob_name, pdf_bytes):
    blob_service = get_blob_service_client()
    if blob_service is None:
        return None

    container_client = blob_service.get_container_client(AZURE_STORAGE_CONTAINER)
    try:
        container_client.create_container()
    except Exception:
        pass

    blob_client = container_client.get_blob_client(blob_name)
    content_settings = ContentSettings(content_type="application/pdf") if ContentSettings else None
    blob_client.upload_blob(pdf_bytes, overwrite=True, content_settings=content_settings)
    return blob_client.url


def get_storage_account_key_from_connection_string():
    if not AZURE_STORAGE_CONNECTION_STRING:
        return ""

    for part in AZURE_STORAGE_CONNECTION_STRING.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        if key.strip().lower() == "accountkey":
            return value.strip()
    return ""


def generate_blob_read_sas_url(blob_name, expiry_minutes):
    blob_service = get_blob_service_client()
    if blob_service is None:
        raise RuntimeError("Could not initialize Azure Blob client.")

    try:
        from azure.storage.blob import generate_blob_sas
        from azure.storage.blob import BlobSasPermissions
    except ImportError as exc:
        raise RuntimeError("SAS generation requires azure-storage-blob package.") from exc

    account_name = str(getattr(blob_service, "account_name", "") or "").strip()
    if not account_name:
        raise RuntimeError("Could not determine Azure storage account name.")

    account_key = get_storage_account_key_from_connection_string()
    if not account_key:
        raise RuntimeError("Account key not found in AZURE_STORAGE_CONNECTION_STRING.")

    safe_minutes = max(1, min(int(expiry_minutes), 7 * 24 * 60))
    expires_at_utc = datetime.now(timezone.utc) + timedelta(minutes=safe_minutes)

    sas_token = generate_blob_sas(
        account_name=account_name,
        container_name=AZURE_STORAGE_CONTAINER,
        blob_name=blob_name,
        account_key=account_key,
        permission=BlobSasPermissions(read=True),
        expiry=expires_at_utc,
    )
    if not sas_token:
        raise RuntimeError("Failed to generate SAS token.")

    base_blob_url = f"https://{account_name}.blob.core.windows.net/{AZURE_STORAGE_CONTAINER}/{blob_name}"
    sas_url = f"{base_blob_url}?{sas_token}"
    return sas_url, expires_at_utc.strftime("%Y-%m-%dT%H:%M:%SZ")


def extract_blob_name_from_url(blob_url):
    if not blob_url:
        return ""

    parsed = urlparse(str(blob_url))
    marker = f"/{AZURE_STORAGE_CONTAINER}/"
    idx = parsed.path.find(marker)
    if idx < 0:
        return ""

    return unquote(parsed.path[idx + len(marker):]).lstrip("/")


def build_ticket_image_view_url(image_url):
    raw_url = str(image_url or "").strip()
    if not raw_url:
        return ""

    blob_name = extract_blob_name_from_url(raw_url)
    if not blob_name:
        return raw_url

    try:
        sas_url, _expires = generate_blob_read_sas_url(blob_name, AZURE_IMAGE_SAS_MINUTES)
        return sas_url
    except Exception as exc:
        app.logger.warning("Could not generate image SAS for '%s': %s", blob_name, exc)
        return raw_url


def upload_download_audit_blob(category, filename, file_bytes, mimetype):
    blob_service = get_blob_service_client()
    if blob_service is None:
        return None

    safe_category = re.sub(r"[^a-zA-Z0-9_\-]", "_", str(category or "download"))
    safe_name = re.sub(r"[^a-zA-Z0-9._\-]", "_", str(filename or "file"))
    stamp = app_now().strftime("%Y/%m/%d/%H%M%S_%f")
    blob_name = (
        f"{AZURE_DOWNLOADS_BLOB_PREFIX}/{safe_category}/{stamp}_{safe_name}"
        if AZURE_DOWNLOADS_BLOB_PREFIX
        else f"{safe_category}/{stamp}_{safe_name}"
    )

    remote_ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "")[:120]
    endpoint = str(request.endpoint or "")[:120]

    metadata = {
        "endpoint": re.sub(r"[^a-zA-Z0-9_\-]", "_", endpoint),
        "remote_ip": re.sub(r"[^a-zA-Z0-9_\-:., ]", "_", remote_ip),
        "downloaded_at": datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ"),
    }

    container_client = blob_service.get_container_client(AZURE_STORAGE_CONTAINER)
    try:
        container_client.create_container()
    except Exception:
        pass

    blob_client = container_client.get_blob_client(blob_name)
    content_settings = ContentSettings(content_type=mimetype) if ContentSettings and mimetype else None
    blob_client.upload_blob(file_bytes, overwrite=True, content_settings=content_settings, metadata=metadata)
    return blob_client.url


def upload_jobs_cache_blob(file_bytes):
    blob_service = get_blob_service_client()
    if blob_service is None or not AZURE_JOBS_CACHE_BLOB_NAME:
        return None

    container_client = blob_service.get_container_client(AZURE_STORAGE_CONTAINER)
    try:
        container_client.create_container()
    except Exception:
        pass

    blob_client = container_client.get_blob_client(AZURE_JOBS_CACHE_BLOB_NAME)
    content_settings = ContentSettings(content_type="text/csv") if ContentSettings else None
    blob_client.upload_blob(file_bytes, overwrite=True, content_settings=content_settings)
    return blob_client.url


def download_jobs_cache_blob():
    blob_service = get_blob_service_client()
    if blob_service is None or not AZURE_JOBS_CACHE_BLOB_NAME:
        return None

    try:
        blob_client = blob_service.get_blob_client(
            container=AZURE_STORAGE_CONTAINER,
            blob=AZURE_JOBS_CACHE_BLOB_NAME,
        )
        content = blob_client.download_blob().readall()
        return content if content else None
    except Exception as exc:
        app.logger.warning("Could not download jobs cache blob '%s': %s", AZURE_JOBS_CACHE_BLOB_NAME, exc)
        return None


def delete_pdf_blob_if_needed(pdf_path):
    if not pdf_path or not str(pdf_path).lower().startswith("http"):
        return False

    blob_service = get_blob_service_client()
    if blob_service is None:
        return False

    parsed = urlparse(str(pdf_path))
    marker = f"/{AZURE_STORAGE_CONTAINER}/"
    idx = parsed.path.find(marker)
    if idx < 0:
        return False

    blob_name = unquote(parsed.path[idx + len(marker):]).lstrip("/")
    if not blob_name:
        return False

    try:
        blob_service.get_blob_client(container=AZURE_STORAGE_CONTAINER, blob=blob_name).delete_blob(delete_snapshots="include")
        return True
    except Exception as exc:
        app.logger.warning("Could not delete blob '%s': %s", blob_name, exc)
        return False


def write_temp_pdf_for_print(pdf_bytes, name_prefix):
    cache_dir = BASE_DIR / "tmp_print"
    cache_dir.mkdir(parents=True, exist_ok=True)
    stamp = app_now().strftime("%Y%m%d_%H%M%S_%f")
    path = cache_dir / f"{name_prefix}_{stamp}.pdf"
    with open(path, "wb") as f:
        f.write(pdf_bytes)
    return str(path)


def format_ticket_datetime(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return text
    return dt.strftime("%m-%d-%Y - %I:%M %p")


def format_ticket_time(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return text
    return dt.strftime("%I:%M %p")


def format_currency(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    return f"${amount:,.2f}"


@app.template_filter("ticket_datetime")
def ticket_datetime_filter(value):
    return format_ticket_datetime(value)


def resolve_jobs_csv_path():
    csv_url = os.getenv("JOBS_CSV_URL", "").strip()
    if csv_url:
        cache_path_raw = os.getenv("JOBS_CSV_CACHE_PATH", "").strip()
        if cache_path_raw:
            cache_path = Path(cache_path_raw)
            if not cache_path.is_absolute():
                cache_path = BASE_DIR / cache_path
        else:
            cache_path = DATA_DIR / "jobs_remote_cache.csv"

        cache_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            download_jobs_csv_from_url(csv_url, cache_path)
            app.logger.info("Jobs CSV downloaded from URL to %s", cache_path)

            try:
                with open(cache_path, "rb") as f:
                    blob_url = upload_jobs_cache_blob(f.read())
                if blob_url:
                    app.logger.info("Jobs CSV uploaded to blob cache at %s", blob_url)
            except Exception as exc:
                app.logger.warning("Jobs CSV upload to blob cache failed: %s", exc)

            return cache_path
        except Exception as exc:
            blob_content = download_jobs_cache_blob()
            if blob_content:
                tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
                with open(tmp_path, "wb") as f:
                    f.write(blob_content)
                tmp_path.replace(cache_path)
                app.logger.warning(
                    "Jobs CSV URL download failed (%s). Using blob cache '%s'.",
                    exc,
                    AZURE_JOBS_CACHE_BLOB_NAME,
                )
                return cache_path

            if cache_path.exists():
                app.logger.warning(
                    "Jobs CSV URL download failed (%s). Using cached file at %s.",
                    exc,
                    cache_path,
                )
                return cache_path
            app.logger.warning("Jobs CSV URL download failed and no cache is available: %s", exc)

    configured = os.getenv("JOBS_CSV_PATH", "").strip()
    candidates = []

    if configured:
        configured_path = Path(configured)
        if not configured_path.is_absolute():
            configured_path = BASE_DIR / configured_path
        candidates.append(configured_path)
    else:
        candidates.append(BASE_DIR / "data" / "jobs.csv")
        candidates.append(Path(r"G:\My Drive\Jobs Master ALL.csv"))

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def normalize_jobs_csv_url(url):
    parsed = urlparse(url)
    host = (parsed.netloc or "").lower()
    if "drive.google.com" not in host:
        return url

    file_id = extract_google_drive_file_id(url)
    if file_id:
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    return url


def extract_google_drive_file_id(url):
    parsed = urlparse(url)
    path = parsed.path or ""

    file_match = re.search(r"/file/d/([^/]+)", path)
    if file_match:
        return file_match.group(1)

    query_id = parse_qs(parsed.query).get("id", [])
    if query_id:
        return query_id[0]

    return None


def download_jobs_csv_from_url(url, destination_path):
    download_url = normalize_jobs_csv_url(url)
    request = Request(
        download_url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; McCrakenTicketSystem/1.0)"
        },
    )

    with urlopen(request, timeout=30) as response:
        content = response.read()

    if not content:
        raise RuntimeError("Downloaded CSV is empty.")

    head = content[:512].lstrip().lower()
    if head.startswith(b"<!doctype html") or head.startswith(b"<html"):
        raise RuntimeError(
            "URL returned HTML instead of CSV. Ensure the Google Drive file is shared and downloadable."
        )

    tmp_path = destination_path.with_suffix(destination_path.suffix + ".tmp")
    with open(tmp_path, "wb") as f:
        f.write(content)
    tmp_path.replace(destination_path)


def create_db_connection():
    database_url = os.getenv("DATABASE_URL", "").strip()
    if database_url:
        conn = psycopg2.connect(
            database_url,
            cursor_factory=RealDictCursor,
            connect_timeout=PG_CONNECT_TIMEOUT,
        )
    else:
        host = os.getenv("PGHOST", "").strip()
        database = os.getenv("PGDATABASE", "").strip()
        user = os.getenv("PGUSER", "").strip()
        password = os.getenv("PGPASSWORD", "").strip()
        missing = []
        if not host:
            missing.append("PGHOST")
        if not database:
            missing.append("PGDATABASE")
        if not user:
            missing.append("PGUSER")
        if not password:
            missing.append("PGPASSWORD")
        if missing:
            raise RuntimeError(
                "Missing PostgreSQL settings: "
                + ", ".join(missing)
                + ". Set DATABASE_URL or populate .env with PGHOST/PGDATABASE/PGUSER/PGPASSWORD."
            )
        conn = psycopg2.connect(
            host=host,
            port=int(os.getenv("PGPORT", "5432").strip()),
            dbname=database,
            user=user,
            password=password,
            sslmode=os.getenv("PGSSLMODE", "require").strip(),
            cursor_factory=RealDictCursor,
            connect_timeout=PG_CONNECT_TIMEOUT,
        )
    conn.autocommit = False
    return conn


def get_db():
    ensure_app_initialized()
    if "db" not in g:
        g.db = create_db_connection()
    return g.db


@app.teardown_appcontext
def close_db(_exception):
    db = g.pop("db", None)
    if db is not None:
        db.close() 


def init_db():
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    with closing(create_db_connection()) as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT to_regclass('public.tickets') AS tickets_table")
            existing = cursor.fetchone() or {}
            has_existing_core_tables = bool(existing.get("tickets_table"))

        # For existing PostgreSQL deployments, skip schema.sql and apply migrations only.
        # This prevents CREATE INDEX errors when schema.sql references newer columns.
        if not has_existing_core_tables:
            with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
                schema_sql = f.read()
            with conn.cursor() as cursor:
                cursor.execute(schema_sql)

        ensure_db_migrations(conn)
        conn.commit()


def ensure_app_initialized():
    global _db_initialized
    if _db_initialized:
        return

    # In managed environments, schema/bootstrap can be run separately to keep worker startup fast.
    if not AUTO_DB_BOOTSTRAP:
        _db_initialized = True
        return

    with _db_init_lock:
        if _db_initialized:
            return
        init_db()
        _db_initialized = True


def ensure_db_migrations(conn):
    with conn.cursor() as cursor:
        cursor.execute("ALTER TABLE jobs_cache ADD COLUMN IF NOT EXISTS tax_exempt TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS customer_snapshot TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS tax_exempt TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS cost REAL NOT NULL DEFAULT 0")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS active BOOLEAN NOT NULL DEFAULT TRUE")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS modified_to_id BIGINT")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS image_url TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS image_status TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS image_error TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS image_generated_at TEXT")
        cursor.execute("ALTER TABLE tickets ALTER COLUMN pdf_path DROP NOT NULL")
        cursor.execute("ALTER TABLE tickets ALTER COLUMN pdf_blob DROP NOT NULL")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tickets_active ON tickets(active)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tickets_image_status ON tickets(image_status)")
        cursor.execute("ALTER TABLE trucks ADD COLUMN IF NOT EXISTS truck_size TEXT NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE trucks ADD COLUMN IF NOT EXISTS hauled_by TEXT NOT NULL DEFAULT ''")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS manual_jobs (
                id BIGSERIAL PRIMARY KEY,
                job_code TEXT NOT NULL,
                job_name TEXT NOT NULL,
                customer TEXT NOT NULL DEFAULT '',
                tax_exempt TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                UNIQUE(job_code, job_name)
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS rfid_notifications (
                id BIGSERIAL PRIMARY KEY,
                event_type TEXT NOT NULL DEFAULT 'known_truck_detected',
                truck_number TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT 'rfid',
                message TEXT NOT NULL,
                detected_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                status TEXT NOT NULL DEFAULT 'pending',
                decided_at TIMESTAMPTZ
            )
            """
        )
        cursor.execute("ALTER TABLE rfid_notifications ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'pending'")
        cursor.execute("ALTER TABLE rfid_notifications ADD COLUMN IF NOT EXISTS decided_at TIMESTAMPTZ")
        cursor.execute(
            """
            UPDATE rfid_notifications
            SET status = 'pending'
            WHERE status IS NULL OR status NOT IN ('pending', 'approved', 'denied')
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS rfid_epc_truck_map (
                id BIGSERIAL PRIMARY KEY,
                epc TEXT NOT NULL UNIQUE,
                truck_number TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT '',
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rfid_epc_map_epc ON rfid_epc_truck_map(LOWER(TRIM(epc)))")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rfid_epc_map_truck ON rfid_epc_truck_map(LOWER(TRIM(truck_number)))")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS rfid_tag_catalog (
                serial_number TEXT PRIMARY KEY,
                epc TEXT NOT NULL UNIQUE,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS rfid_truck_tag_assignment (
                id BIGSERIAL PRIMARY KEY,
                serial_number TEXT NOT NULL UNIQUE REFERENCES rfid_tag_catalog(serial_number) ON DELETE RESTRICT,
                truck_id BIGINT NOT NULL UNIQUE REFERENCES trucks_main(id) ON DELETE RESTRICT,
                notes TEXT NOT NULL DEFAULT '',
                assigned_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rfid_tag_catalog_epc ON rfid_tag_catalog(LOWER(TRIM(epc)))")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rfid_assign_serial ON rfid_truck_tag_assignment(LOWER(TRIM(serial_number)))")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rfid_assign_truck ON rfid_truck_tag_assignment(truck_id)")


def refresh_jobs_on_startup():
    auto_refresh = os.getenv("AUTO_REFRESH_JOBS_ON_STARTUP", "1").strip().lower()
    if auto_refresh not in {"1", "true", "yes", "on"}:
        app.logger.info("Startup jobs refresh disabled by AUTO_REFRESH_JOBS_ON_STARTUP.")
        return

    with app.app_context():
        db = get_db()
        try:
            count = refresh_jobs_cache(db)
            db.commit()
            app.logger.info("Startup jobs refresh complete. %s rows synced.", count)
        except Exception as exc:
            db.rollback()
            app.logger.warning("Startup jobs refresh skipped/failed: %s", exc)


def next_ticket_number(db):
    year = app_now().year
    with db.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO ticket_sequence (ticket_year, last_value)
            VALUES (%s, 0)
            ON CONFLICT (ticket_year) DO NOTHING
            """,
            (year,),
        )
        cursor.execute(
            """
            UPDATE ticket_sequence
            SET last_value = last_value + 1
            WHERE ticket_year = %s
            RETURNING last_value
            """,
            (year,),
        )
        row = cursor.fetchone()
    next_value = int(row["last_value"])
    ticket_number = f"DT-{year}-{next_value:06d}"
    return ticket_number, year, next_value


def truck_size_to_axle_index(truck_size):
    size_text = str(truck_size or "").strip().lower()
    if not size_text:
        return None

    match = re.search(r"axle\s*([0-9]+(?:\.[0-9]+)?)", size_text)
    if not match:
        return None

    axle_value = float(match.group(1))
    axle_index = int(axle_value)
    if axle_value != axle_index:
        return None
    return axle_index if 1 <= axle_index <= 9 else None


def calculate_ticket_cost(truck, material, quantity_num):
    # print(f"Calculating cost for truck: {truck}, material: {material}, quantity: {quantity_num}")
    if not truck or not material:
        return 0.0
    track_size_map = { 
        "Axle 1":"axle_1", "Tandem":"tandem","TriAxle":"triaxle","Axle 4_5":"axle_4_5","Axle 6":"axle_6","Semi":"semi","HydVac":"hydvac","Hydrovac":"hydvac","DIRT_IN":"dirt_in"
    }
    price_per_load = float(material.get(track_size_map.get(truck.get("truck_size"))) or 0)
    return round(price_per_load * quantity_num, 2)




def to_pdf_bytes(ticket):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    company_header_raw = os.getenv("COMPANY_HEADER", "").strip()
    if company_header_raw:
        company_header_lines = [line.strip() for line in company_header_raw.split("|") if line.strip()]
    else:
        company_header_lines = [
            "McCracken Materials, LLC",
            "13675 McCracken Road",
            "Garfield Heights, Ohio 44125",
            "Phone: (216) 206-2600",
        ]

    def draw_field(label, value, x, y, w, h, label_width=70, font_size=8):
        pdf.rect(x, y - h, w, h)
        pdf.setFont("Helvetica-Bold", font_size)
        pdf.drawString(x + 4, y - 12, str(label))
        pdf.line(x + label_width, y - h, x + label_width, y)
        pdf.setFont("Helvetica", font_size)
        pdf.drawString(x + label_width + 4, y - 12, str(value or ""))

    def draw_ticket_section(copy_title, include_signature_line, section_top, section_bottom):
        left = 36
        right = width - 36
        box_w = right - left

        section_h = section_top - section_bottom

        pdf.setLineWidth(1)
        pdf.rect(left, section_bottom + 10, box_w, section_h - 20)

        company_start_y = section_top - 22

        for index, line in enumerate(company_header_lines):
            if index == 0:
                pdf.setFont("Helvetica-Bold", 9)
            else:
                pdf.setFont("Helvetica", 7)

            pdf.drawCentredString(width / 2, company_start_y - (index * 10), line)

        y_top = section_top - 72
        ticket_type ="DUMP TICKET" if ticket["direction"] == "IN" else "MATERIAL TICKET"

        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawCentredString(width / 2, y_top, ticket_type)

        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawCentredString(width / 2, y_top - 14, copy_title.upper())

        draw_field(
            "Ticket #",
            ticket["ticket_number"],
            left + 10,
            y_top - 26,
            250,
            18,
            label_width=55,
            font_size=8
        )

        draw_field(
            "Date/Time",
            format_ticket_datetime(ticket["created_at"]),
            left + 270,
            y_top - 26,
            box_w - 280,
            18,
            label_width=62,
            font_size=8
        )

        direction_text = "IN  [X]   OUT [ ]" if ticket["direction"] == "IN" else "IN  [ ]   OUT [X]"

        draw_field(
            "Direction",
            direction_text,
            left + 10,
            y_top - 48,
            box_w - 20,
            18,
            label_width=62,
            font_size=8
        )

        draw_field(
            "Job #",
            ticket["job_code_snapshot"],
            left + 10,
            y_top - 70,
            180,
            18,
            label_width=42,
            font_size=8
        )

        draw_field(
            "Job Name",
            ticket["job_name_snapshot"],
            left + 195,
            y_top - 70,
            box_w - 205,
            18,
            label_width=58,
            font_size=8
        )

        draw_field(
            "Customer",
            ticket.get("customer_snapshot", ""),
            left + 10,
            y_top - 92,
            box_w - 20,
            18,
            label_width=58,
            font_size=8
        )

        draw_field(
            "Truck #",
            ticket["truck_number_snapshot"],
            left + 10,
            y_top - 114,
            180,
            18,
            label_width=50,
            font_size=8
        )

        draw_field(
            "Material",
            ticket["material_name_snapshot"],
            left + 195,
            y_top - 114,
            box_w - 205,
            18,
            label_width=52,
            font_size=8
        )

        draw_field(
            "Quantity",
            f"{ticket['quantity']}",
            left + 10,
            y_top - 136,
            180,
            18,
            label_width=55,
            font_size=8
        )

        draw_field(
            "Unit",
            ticket["unit"],
            left + 195,
            y_top - 136,
            120,
            18,
            label_width=32,
            font_size=8
        )

        # draw_field(
        #     "Cost",
        #     f"${float(ticket.get('cost', 0) or 0):.2f}",
        #     left + 320,
        #     y_top - 136,
        #     box_w - 330,
        #     18,
        #     label_width=34,
        #     font_size=8
        # )

        notes_y = y_top - 158
        notes_h = 42

        pdf.rect(left + 10, notes_y - notes_h, box_w - 20, notes_h)

        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(left + 16, notes_y - 12, "Notes")

        pdf.setFont("Helvetica", 8)
        note_text = ticket.get("notes", "") or ""

        if len(note_text) > 120:
            note_text = note_text[:117] + "..."

        pdf.drawString(left + 58, notes_y - 12, note_text)

        if include_signature_line:
            sig_y = notes_y - notes_h - 12

            draw_field(
                "Driver Name",
                "",
                left + 10,
                sig_y,
                (box_w - 30) / 2,
                18,
                label_width=62,
                font_size=8
            )

            draw_field(
                "Signature",
                "",
                left + 20 + (box_w - 30) / 2,
                sig_y,
                (box_w - 30) / 2,
                18,
                label_width=58,
                font_size=8
            )

        pdf.setFont("Helvetica", 7)
        pdf.drawRightString(
            right - 8,
            section_bottom + 18,
            f"Printed: {format_ticket_datetime(app_now())}"
        )

    middle_y = height / 2

    draw_ticket_section(
        "Driver Copy Signature Required",
        True,
        section_top=height,
        section_bottom=middle_y
    )

    draw_ticket_section(
        "Internal Billing Copy",
        False,
        section_top=middle_y,
        section_bottom=0
    )

    pdf.setDash(3, 3)
    pdf.line(36, middle_y, width - 36, middle_y)
    pdf.setDash()

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    return buffer.read()

def report_to_pdf_bytes(tickets, totals_by_unit, totals_by_material, filters):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    bold = styles["Heading3"]

    elements = []

    # ---- Title ----
    elements.append(Paragraph("Ticket Report", bold))
    elements.append(
        Paragraph(
            f"Generated: {app_now().strftime('%m-%d-%Y %I:%M %p')}",
            normal,
        )
    )
    elements.append(Spacer(1, 12))

    # ---- Tickets Table ----
    table_data = [
        ["Ticket #", "Date/Time", "Customer", "Dir", "Material", "Qty", "Unit", "Cost"]
    ]

    for t in tickets:
        table_data.append([
            Paragraph(str(t["ticket_number"]), normal),
            Paragraph(format_ticket_datetime(t["created_at"]), normal),
            Paragraph(t["customer_snapshot"] or "", normal),
            Paragraph(t["direction"], normal),
            Paragraph(t["material_name_snapshot"], normal),
            Paragraph(f"{t['quantity']:.2f}", normal),
            Paragraph(t["unit"], normal),
            Paragraph(f"${float(t.get('cost', 0) or 0):.2f}", normal),
        ])

    ticket_table = Table(
        table_data,
        colWidths=[65, 85, 110, 30, 90, 40, 35, 55],
        repeatRows=1
    )

    ticket_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("ALIGN", (7, 1), (7, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elements.append(ticket_table)
    elements.append(Spacer(1, 20))

    # ---- Totals By Unit ----
    elements.append(Paragraph("Totals By Unit", bold))
    totals_unit_table = Table(
        [["Unit", "Total Quantity", "Total Cost"]] +
        [[r["unit"], f"{r['total_quantity']:.2f}", f"${float(r.get('total_cost', 0) or 0):.2f}"] for r in totals_by_unit],
        colWidths=[90, 110, 110],
    )
    totals_unit_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(totals_unit_table)
    elements.append(Spacer(1, 20))

    # ---- Totals By Material ----
    elements.append(Paragraph("Totals By Material", bold))
    totals_mat_table = Table(
        [["Material", "Unit", "Total Quantity", "Total Cost"]] +
        [
            [
                r["material_name_snapshot"],
                r["unit"],
                f"{r['total_quantity']:.2f}",
                f"${float(r.get('total_cost', 0) or 0):.2f}",
            ]
            for r in totals_by_material
        ],
        colWidths=[180, 70, 100, 100],
    )
    totals_mat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(totals_mat_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def fetch_credit_card_sales_rows(db, report_date, customer_match):
    like_value = f"%{customer_match}%"
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                t.ticket_number,
                t.created_at,
                t.customer_snapshot,
                t.job_code_snapshot,
                t.job_name_snapshot,
                t.truck_number_snapshot,
                t.material_name_snapshot,
                t.quantity,
                t.unit,
                t.cost
            FROM tickets t
            WHERE COALESCE(t.active, TRUE) = TRUE
              AND date(t.created_at) = date(%s)
              AND COALESCE(t.customer_snapshot, '') ILIKE %s
            ORDER BY t.created_at ASC, t.id ASC
            """,
            (report_date.isoformat(), like_value),
        )
        return cursor.fetchall()


def fetch_non_credit_card_sales_rows(db, start_date, end_date, customer_match):
    like_value = f"%{customer_match}%"
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                t.ticket_number,
                t.created_at,
                t.customer_snapshot,
                t.job_code_snapshot,
                t.job_name_snapshot,
                t.truck_number_snapshot,
                t.material_name_snapshot,
                t.quantity,
                t.unit,
                t.cost
            FROM tickets t
            WHERE COALESCE(t.active, TRUE) = TRUE
                            AND date(t.created_at) BETWEEN date(%s) AND date(%s)
              AND COALESCE(t.customer_snapshot, '') NOT ILIKE %s
            ORDER BY COALESCE(NULLIF(TRIM(t.customer_snapshot), ''), 'zzzzzz') ASC, t.created_at ASC, t.id ASC
            """,
                        (start_date.isoformat(), end_date.isoformat(), like_value),
        )
        return cursor.fetchall()


def credit_card_daily_report_to_pdf_bytes(rows, report_date, customer_match):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    heading = styles["Heading2"]
    subheading = styles["Heading3"]

    elements = []
    elements.append(Paragraph("Daily Credit Card Sales Report", heading))
    elements.append(Paragraph(f"Report Date: {report_date.strftime('%m/%d/%Y')}", normal))
    elements.append(Paragraph(f"Customer match: {customer_match}", normal))
    elements.append(Paragraph(f"Generated: {app_now().strftime('%m/%d/%Y %I:%M %p %Z')}", normal))
    elements.append(Spacer(1, 12))

    if not rows:
        elements.append(Paragraph("No credit card sales for this day.", subheading))
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    total_amount = sum(float(r.get("cost") or 0) for r in rows)
    total_count = len(rows)

    col_widths = [85, 100, 100, 150, 80, 140, 65]

    def clip_to_width(value, width_points, font_name="Helvetica", font_size=9):
        text = str(value or "")
        if not text:
            return ""

        usable_width = max(8, float(width_points) - 8)
        if stringWidth(text, font_name, font_size) <= usable_width:
            return text

        while text and stringWidth(text, font_name, font_size) > usable_width:
            text = text[:-1]
        return text

    elements.append(Paragraph(f"Total transactions: {total_count}", normal))
    elements.append(Paragraph(f"Total amount: {format_currency(total_amount)}", normal))
    elements.append(Spacer(1, 10))

    table_data = [[
        "Ticket #",
        "Date/Time",
        "Customer",
        "Job",
        "Truck",
        "Material",
        "Cost",
    ]]

    for row in rows:
        job_text = " - ".join(
            part for part in [
                str(row.get("job_code_snapshot") or "").strip(),
                str(row.get("job_name_snapshot") or "").strip(),
            ]
            if part
        )
        table_data.append([
            clip_to_width(str(row.get("ticket_number") or "").strip() or "-", col_widths[0]),
            clip_to_width(format_ticket_datetime(row.get("created_at")), col_widths[1]),
            clip_to_width(str(row.get("customer_snapshot") or ""), col_widths[2]),
            clip_to_width(job_text, col_widths[3]),
            clip_to_width(str(row.get("truck_number_snapshot") or ""), col_widths[4]),
            clip_to_width(str(row.get("material_name_snapshot") or ""), col_widths[5]),
            format_currency(row.get("cost")),
        ])

    table = Table(
        table_data,
        colWidths=col_widths,
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def non_credit_card_daily_report_to_pdf_bytes(rows, report_start_date, report_end_date, customer_match):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    heading = styles["Heading2"]
    subheading = styles["Heading3"]

    elements = []
    elements.append(Paragraph("Daily Non-Credit Card Sales Report", heading))
    elements.append(
        Paragraph(
            f"Report Range: {report_start_date.strftime('%m/%d/%Y')} - {report_end_date.strftime('%m/%d/%Y')}",
            normal,
        )
    )
    elements.append(Paragraph(f"Excluding customer match: {customer_match}", normal))
    elements.append(Paragraph(f"Generated: {app_now().strftime('%m/%d/%Y %I:%M %p %Z')}", normal))
    elements.append(Spacer(1, 12))

    if not rows:
        elements.append(Paragraph("No non-credit-card sales for this day.", subheading))
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    total_amount = sum(float(r.get("cost") or 0) for r in rows)
    total_count = len(rows)

    def normalize_customer_name(value):
        return " ".join(str(value or "").strip().lower().split())

    our_customers = {"mrex", "petty group llc", "redcon"}
    external_grouped_rows = {}
    internal_grouped_rows = {}

    for row in rows:
        customer_name = str(row.get("customer_snapshot") or "").strip() or "(No Customer)"
        customer_key = normalize_customer_name(customer_name)

        if customer_key in our_customers:
            job_code = str(row.get("job_code_snapshot") or "").strip()
            job_name = str(row.get("job_name_snapshot") or "").strip()
            job_key = (job_code, job_name)
            internal_grouped_rows.setdefault(customer_name, {}).setdefault(job_key, []).append(row)
        else:
            external_grouped_rows.setdefault(customer_name, []).append(row)

    elements.append(Paragraph(f"Total transactions: {total_count}", normal))
    elements.append(Paragraph(f"Total amount: {format_currency(total_amount)}", normal))
    elements.append(Paragraph(f"Total customers: {len(external_grouped_rows) + len(internal_grouped_rows)}", normal))
    elements.append(Spacer(1, 10))

    col_widths = [130, 95, 220, 30, 180, 65]

    def clip_to_width(value, width_points, font_name="Helvetica", font_size=9):
        text = str(value or "")
        if not text:
            return ""

        # Keep a little padding so clipped text never touches or spills grid lines.
        usable_width = max(8, float(width_points) - 8)
        if stringWidth(text, font_name, font_size) <= usable_width:
            return text

        while text and stringWidth(text, font_name, font_size) > usable_width:
            text = text[:-1]
        return text

    # Section 1: customers that are NOT MREX/Petty Group/Redcon.
    # Each customer starts on a new page with one table for that customer.
    external_customers = sorted(external_grouped_rows.keys(), key=lambda x: x.lower())
    first_detail_page = True

    def append_external_customer_page(customer_name, customer_rows):
        customer_total = sum(float(r.get("cost") or 0) for r in customer_rows)

        elements.append(Paragraph("Section 1: Non-MREX / Non-Petty Group / Non-Redcon", subheading))
        elements.append(Paragraph(f"Customer: {customer_name}", subheading))
        elements.append(Paragraph(f"Transactions: {len(customer_rows)} | Total amount: {format_currency(customer_total)}", normal))
        elements.append(Spacer(1, 6))

        table_data = [["Date/Time", "Ticket #", "Job", "Truck", "Material", "Cost"]]
        for row in customer_rows:
            job_text = " - ".join(
                part
                for part in [
                    str(row.get("job_code_snapshot") or "").strip(),
                    str(row.get("job_name_snapshot") or "").strip(),
                ]
                if part
            )
            table_data.append([
                clip_to_width(format_ticket_datetime(row.get("created_at")), col_widths[0]),
                clip_to_width(str(row.get("ticket_number") or "").strip() or "-", col_widths[1]),
                clip_to_width(job_text, col_widths[2]),
                clip_to_width(str(row.get("truck_number_snapshot") or ""), col_widths[3]),
                clip_to_width(str(row.get("material_name_snapshot") or ""), col_widths[4]),
                format_currency(row.get("cost")),
            ])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (5, 1), (5, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)

    for customer_name in external_customers:
        if first_detail_page:
            first_detail_page = False
        else:
            elements.append(PageBreak())
        append_external_customer_page(customer_name, external_grouped_rows[customer_name])

    # Section 2: our customers (MREX/Petty Group/Redcon), grouped by job.
    # Each job starts on a new page.
    internal_customers = sorted(internal_grouped_rows.keys(), key=lambda x: x.lower())

    def append_internal_job_page(customer_name, job_key, job_rows):
        job_code, job_name = job_key
        job_code = job_code or "(No Job Code)"
        job_name = job_name or "(No Job Name)"
        job_total = sum(float(r.get("cost") or 0) for r in job_rows)

        elements.append(Paragraph("Section 2: MREX / Petty Group / Redcon (Grouped by Job)", subheading))
        elements.append(Paragraph(f"Customer: {customer_name}", subheading))
        elements.append(Paragraph(f"Job: {job_code} - {job_name}", subheading))
        elements.append(Paragraph(f"Transactions: {len(job_rows)} | Total amount: {format_currency(job_total)}", normal))
        elements.append(Spacer(1, 6))

        job_col_widths = [95, 110, 95, 230, 70]
        table_data = [["Date/Time", "Ticket #", "Truck", "Material", "Cost"]]
        for row in job_rows:
            table_data.append([
                clip_to_width(format_ticket_datetime(row.get("created_at")), job_col_widths[0]),
                clip_to_width(str(row.get("ticket_number") or "").strip() or "-", job_col_widths[1]),
                clip_to_width(str(row.get("truck_number_snapshot") or ""), job_col_widths[2]),
                clip_to_width(str(row.get("material_name_snapshot") or ""), job_col_widths[3]),
                format_currency(row.get("cost")),
            ])

        table = Table(table_data, colWidths=job_col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)

    for customer_name in internal_customers:
        job_groups = internal_grouped_rows[customer_name]
        sorted_jobs = sorted(job_groups.keys(), key=lambda key: ((key[0] or "").lower(), (key[1] or "").lower()))
        for job_key in sorted_jobs:
            if first_detail_page:
                first_detail_page = False
            else:
                elements.append(PageBreak())
            append_internal_job_page(customer_name, job_key, job_groups[job_key])

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def customer_grouped_report_to_pdf_bytes(rows, filters):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    heading = styles["Heading2"]
    subheading = styles["Heading3"]

    def _v(value):
        return str(value or "").strip()

    date_from = _v(filters.get("date_from"))
    date_to = _v(filters.get("date_to"))
    direction = _v(filters.get("direction"))
    job_label = _v(filters.get("job_label"))
    customer = _v(filters.get("customer"))
    material_label = _v(filters.get("material_label"))

    elements = []
    elements.append(Paragraph("Customer Grouped Sales Report", heading))
    elements.append(Paragraph(f"Generated: {app_now().strftime('%m/%d/%Y %I:%M %p %Z')}", normal))

    applied_filters = []
    if date_from or date_to:
        applied_filters.append(f"Date: {date_from or 'Any'} to {date_to or 'Any'}")
    if direction:
        applied_filters.append(f"Direction: {direction}")
    if job_label:
        applied_filters.append(f"Job: {job_label}")
    if customer:
        applied_filters.append(f"Customer contains: {customer}")
    if material_label:
        applied_filters.append(f"Material: {material_label}")

    if applied_filters:
        for filter_line in applied_filters:
            elements.append(Paragraph(filter_line, normal))
    else:
        elements.append(Paragraph("Filters: None (all active tickets)", normal))

    elements.append(Spacer(1, 12))

    if not rows:
        elements.append(Paragraph("No matching tickets for selected filters.", subheading))
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    total_amount = sum(float(r.get("cost") or 0) for r in rows)
    total_count = len(rows)

    def normalize_customer_name(value):
        return " ".join(str(value or "").strip().lower().split())

    our_customers = {"mrex", "petty group llc", "redcon"}
    external_grouped_rows = {}
    internal_grouped_rows = {}

    for row in rows:
        customer_name = str(row.get("customer_snapshot") or "").strip() or "(No Customer)"
        customer_key = normalize_customer_name(customer_name)

        if customer_key in our_customers:
            job_code = str(row.get("job_code_snapshot") or "").strip()
            job_name = str(row.get("job_name_snapshot") or "").strip()
            job_key = (job_code, job_name)
            internal_grouped_rows.setdefault(customer_name, {}).setdefault(job_key, []).append(row)
        else:
            external_grouped_rows.setdefault(customer_name, []).append(row)

    elements.append(Paragraph(f"Total transactions: {total_count}", normal))
    elements.append(Paragraph(f"Total amount: {format_currency(total_amount)}", normal))
    elements.append(Paragraph(f"Total customers: {len(external_grouped_rows) + len(internal_grouped_rows)}", normal))
    elements.append(Spacer(1, 10))

    col_widths = [110, 95, 220, 50, 180, 65]

    def clip_to_width(value, width_points, font_name="Helvetica", font_size=9):
        text = str(value or "")
        if not text:
            return ""

        usable_width = max(8, float(width_points) - 8)
        if stringWidth(text, font_name, font_size) <= usable_width:
            return text

        while text and stringWidth(text, font_name, font_size) > usable_width:
            text = text[:-1]
        return text

    external_customers = sorted(external_grouped_rows.keys(), key=lambda x: x.lower())
    first_detail_page = True

    def append_external_customer_page(customer_name, customer_rows):
        customer_total = sum(float(r.get("cost") or 0) for r in customer_rows)

        elements.append(Paragraph("Section 1: Non-MREX / Non-Petty Group / Non-Redcon", subheading))
        elements.append(Paragraph(f"Customer: {customer_name}", subheading))
        elements.append(Paragraph(f"Transactions: {len(customer_rows)} | Total amount: {format_currency(customer_total)}", normal))
        elements.append(Spacer(1, 6))

        table_data = [["Date/Time", "Ticket #", "Job", "Truck", "Material", "Cost"]]
        for row in customer_rows:
            job_text = " - ".join(
                part
                for part in [
                    str(row.get("job_code_snapshot") or "").strip(),
                    str(row.get("job_name_snapshot") or "").strip(),
                ]
                if part
            )
            table_data.append([
                clip_to_width(format_ticket_datetime(row.get("created_at")), col_widths[0]),
                clip_to_width(str(row.get("ticket_number") or "").strip() or "-", col_widths[1]),
                clip_to_width(job_text, col_widths[2]),
                clip_to_width(str(row.get("truck_number_snapshot") or ""), col_widths[3]),
                clip_to_width(str(row.get("material_name_snapshot") or ""), col_widths[4]),
                format_currency(row.get("cost")),
            ])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (5, 1), (5, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)

    for customer_name in external_customers:
        if first_detail_page:
            first_detail_page = False
        else:
            elements.append(PageBreak())
        append_external_customer_page(customer_name, external_grouped_rows[customer_name])

    internal_customers = sorted(internal_grouped_rows.keys(), key=lambda x: x.lower())

    def append_internal_job_page(customer_name, job_key, job_rows):
        job_code, job_name = job_key
        job_code = job_code or "(No Job Code)"
        job_name = job_name or "(No Job Name)"
        job_total = sum(float(r.get("cost") or 0) for r in job_rows)

        elements.append(Paragraph("Section 2: MREX / Petty Group / Redcon (Grouped by Job)", subheading))
        elements.append(Paragraph(f"Customer: {customer_name}", subheading))
        elements.append(Paragraph(f"Job: {job_code} - {job_name}", subheading))
        elements.append(Paragraph(f"Transactions: {len(job_rows)} | Total amount: {format_currency(job_total)}", normal))
        elements.append(Spacer(1, 6))

        job_col_widths = [95, 110, 95, 230, 70]
        table_data = [["Date/Time", "Ticket #", "Truck", "Material", "Cost"]]
        for row in job_rows:
            table_data.append([
                clip_to_width(format_ticket_datetime(row.get("created_at")), job_col_widths[0]),
                clip_to_width(str(row.get("ticket_number") or "").strip() or "-", job_col_widths[1]),
                clip_to_width(str(row.get("truck_number_snapshot") or ""), job_col_widths[2]),
                clip_to_width(str(row.get("material_name_snapshot") or ""), job_col_widths[3]),
                format_currency(row.get("cost")),
            ])

        table = Table(table_data, colWidths=job_col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)

    for customer_name in internal_customers:
        job_groups = internal_grouped_rows[customer_name]
        sorted_jobs = sorted(job_groups.keys(), key=lambda key: ((key[0] or "").lower(), (key[1] or "").lower()))
        for job_key in sorted_jobs:
            if first_detail_page:
                first_detail_page = False
            else:
                elements.append(PageBreak())
            append_internal_job_page(customer_name, job_key, job_groups[job_key])

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def daily_report_to_pdf_bytes(job_blocks, report_date, totals):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    left = 18
    right = width - 18
    top = height - 28
    grid_bottom = 95

    cols = 3
    rows = 6
    cell_w = (right - left) / cols
    cell_h = (top - grid_bottom) / rows
    header_h = 24

    header_colors = [
        colors.HexColor("#7dd3fc"),
        colors.HexColor("#86efac"),
        colors.HexColor("#fde68a"),
        colors.HexColor("#f9a8d4"),
    ]

    def clip_text(text, font_name, font_size, max_width):
        value = str(text or "")
        if pdf.stringWidth(value, font_name, font_size) <= max_width:
            return value
        suffix = "..."
        while value and pdf.stringWidth(value + suffix, font_name, font_size) > max_width:
            value = value[:-1]
        return (value + suffix) if value else suffix

    # Keep every truck visible by splitting large jobs across continuation boxes/pages.
    line_height = 12
    text_start_offset = header_h + 12
    text_bottom_padding = 7
    trucks_per_box = max(1, int((cell_h - text_start_offset - text_bottom_padding) // line_height) + 1)

    expanded_blocks = []
    for block in job_blocks:
        trucks = list(block.get("trucks") or [])
        if not trucks:
            expanded_blocks.append(
                {
                    "job_code": block.get("job_code", ""),
                    "job_name": block.get("job_name", ""),
                    "customer": block.get("customer", ""),
                    "trucks": [],
                    "continued": False,
                    "has_more": False,
                }
            )
            continue

        for start in range(0, len(trucks), trucks_per_box):
            segment = trucks[start:start + trucks_per_box]
            expanded_blocks.append(
                {
                    "job_code": block.get("job_code", ""),
                    "job_name": block.get("job_name", ""),
                    "customer": block.get("customer", ""),
                    "trucks": segment,
                    "continued": start > 0,
                    "has_more": (start + trucks_per_box) < len(trucks),
                }
            )

    if not expanded_blocks:
        expanded_blocks = [
            {
                "job_code": "",
                "job_name": "",
                "customer": "",
                "trucks": [],
                "continued": False,
                "has_more": False,
            }
        ]

    max_blocks = cols * rows
    total_pages = (len(expanded_blocks) + max_blocks - 1) // max_blocks

    for page_idx in range(total_pages):
        page_start = page_idx * max_blocks
        page_blocks = expanded_blocks[page_start:page_start + max_blocks]

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(width / 2, height - 16, "Daily Dispatch Report")
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(right, height - 16, f"Page {page_idx + 1}/{total_pages}")

        for idx in range(max_blocks):
            col = idx % cols
            row = idx // cols

            x = left + (col * cell_w)
            y_top = top - (row * cell_h)
            y_bottom = y_top - cell_h

            pdf.setLineWidth(1)
            pdf.setStrokeColor(colors.black)
            pdf.rect(x, y_bottom, cell_w, cell_h)

            if idx >= len(page_blocks):
                continue

            block = page_blocks[idx]
            color = header_colors[(page_start + idx) % len(header_colors)]

            pdf.setFillColor(color)
            pdf.rect(x + 1, y_top - header_h - 1, cell_w - 2, header_h, stroke=0, fill=1)
            pdf.setFillColor(colors.black)

            job_code = str(block.get("job_code") or "").strip()
            job_name = str(block.get("job_name") or "").strip()
            customer = str(block.get("customer") or "").strip()
            if job_code and job_name:
                header_text = f"{job_code} - {job_name}"
            elif job_code:
                header_text = job_code
            elif job_name:
                header_text = job_name
            else:
                header_text = "Unassigned Job"

            if block.get("continued"):
                header_text = f"{header_text} (cont.)"

            pdf.setFont("Helvetica-Bold", 10)
            header_text = clip_text(header_text, "Helvetica-Bold", 10, cell_w - 8)
            pdf.drawString(x + 4, y_top - 11, header_text)

            if customer:
                pdf.setFont("Helvetica", 8)
                customer_text = clip_text(customer, "Helvetica", 8, cell_w - 8)
                pdf.drawString(x + 4, y_top - 20, customer_text)

            y_text = y_top - text_start_offset
            pdf.setFont("Helvetica", 10)

            if not block.get("trucks"):
                pdf.drawString(x + 4, y_text, "- No trucks")
            else:
                for truck_row in block.get("trucks", []):
                    if y_text < y_bottom + text_bottom_padding:
                        break
                    truck_name = str(truck_row.get("truck") or "").strip() or "Unknown Truck"
                    loads = int(truck_row.get("loads") or 0)
                    if loads > 1:
                        line = f"- {truck_name} ({loads})"
                    else:
                        line = f"- {truck_name}"
                    line = clip_text(line, "Helvetica", 10, cell_w - 10)
                    pdf.drawString(x + 4, y_text, line)
                    y_text -= line_height

                if block.get("has_more") and y_text >= y_bottom + text_bottom_padding:
                    pdf.setFont("Helvetica-Oblique", 8)
                    pdf.drawString(x + 4, y_bottom + 4, "continues...")

        date_text = report_date.strftime("%m/%d/%y")
        pdf.setFont("Helvetica-Bold", 20)
        pdf.drawRightString(right - 8, grid_bottom - 12, date_text)

        stats_x = left + (2 * cell_w) + 6
        stats_y = 66
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(stats_x, stats_y, f"TOTAL TRUCKS: {int(totals.get('total_trucks') or 0)}")
        pdf.drawString(stats_x, stats_y - 16, f"TOTAL LOADS IN: {int(totals.get('loads_in') or 0)}")
        pdf.drawString(stats_x, stats_y - 32, f"TOTAL LOADS OUT: {int(totals.get('loads_out') or 0)}")
        pdf.drawString(stats_x, stats_y - 54, f"TOTAL LOADS: {int(totals.get('total_loads') or 0)}")

        pdf.showPage()

    pdf.save()
    buffer.seek(0)
    return buffer.read()


def build_daily_report_data(db, report_date_str):
    def customer_to_color(customer_name):
        value = str(customer_name or "").strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", " ", value)

        # Same customer family should always map to same color.
        if "mrex" in value or "mr ex" in normalized:
            return "#7dd3fc"  # Blue
        if "petty" in value or "redcon" in value:
            return "#86efac"  # Green
        return "#fde68a"  # Yellow

    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                COALESCE(NULLIF(TRIM(t.job_code_snapshot), ''), '') AS job_code,
                COALESCE(NULLIF(TRIM(t.job_name_snapshot), ''), '') AS job_name,
                                COALESCE(NULLIF(TRIM(t.customer_snapshot), ''), '') AS customer,
                COALESCE(NULLIF(TRIM(t.truck_number_snapshot), ''), '') AS truck_number,
                COUNT(*) AS load_count
            FROM tickets t
            WHERE COALESCE(t.active, TRUE) = TRUE
              AND date(t.created_at) = date(%s)
                        GROUP BY 1, 2, 3, 4
                        ORDER BY 1, 2, 3, 4
            """,
            (report_date_str,),
        )
        grouped_rows = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                COUNT(DISTINCT NULLIF(TRIM(t.truck_number_snapshot), '')) AS total_trucks,
                COUNT(*) FILTER (WHERE t.direction = 'IN') AS loads_in,
                COUNT(*) FILTER (WHERE t.direction = 'OUT') AS loads_out,
                COUNT(*) AS total_loads
            FROM tickets t
            WHERE COALESCE(t.active, TRUE) = TRUE
              AND date(t.created_at) = date(%s)
            """,
            (report_date_str,),
        )
        totals = cursor.fetchone() or {
            "total_trucks": 0,
            "loads_in": 0,
            "loads_out": 0,
            "total_loads": 0,
        }

    job_map = {}
    for row in grouped_rows:
        job_code = str(row.get("job_code") or "").strip()
        job_name = str(row.get("job_name") or "").strip()
        customer = str(row.get("customer") or "").strip()
        truck_number = str(row.get("truck_number") or "").strip()
        load_count = int(row.get("load_count") or 0)

        key = (job_code, job_name, customer)
        if key not in job_map:
            job_map[key] = {
                "job_code": job_code,
                "job_name": job_name,
                "customer": customer,
                "customer_color": customer_to_color(customer),
                "trucks": [],
            }

        if truck_number:
            job_map[key]["trucks"].append({
                "truck": truck_number,
                "loads": load_count,
            })

    job_blocks = sorted(
        job_map.values(),
        key=lambda b: (
            -len(b.get("trucks") or []),
            -sum(int(t.get("loads") or 0) for t in (b.get("trucks") or [])),
            str(b.get("job_code") or ""),
            str(b.get("job_name") or ""),
            str(b.get("customer") or ""),
        ),
    )
    return job_blocks, totals


def materials_report_to_pdf_bytes(materials):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    bold = styles["Heading3"]

    elements = []
    elements.append(Paragraph("Materials Export", bold))
    elements.append(
        Paragraph(
            f"Generated: {app_now().strftime('%m-%d-%Y %I:%M %p')}",
            normal,
        )
    )
    elements.append(Spacer(1, 12))

    table_data = [[
        "Material", "Direction", "Price per Cubic Yard",
        "Axle 1", "Tandem", "TriAxle", "Axle 4-5", "Axle 6", "Semi", "HydVac", "Dirt In",
        "Status",
    ]]

    for m in materials:
        table_data.append([
            Paragraph(str(m["material_name"] or ""), normal),
            Paragraph(str(m["direction"] or ""), normal),
            Paragraph(f"${float(m['cost_per_cy']):.2f}" if m["cost_per_cy"] is not None else "", normal),
            Paragraph(f"{float(m['axle_1']):.2f}" if m["axle_1"] is not None else "", normal),
            Paragraph(f"{float(m['tandem']):.2f}" if m["tandem"] is not None else "", normal),
            Paragraph(f"{float(m['triaxle']):.2f}" if m["triaxle"] is not None else "", normal),
            Paragraph(f"{float(m['axle_4_5']):.2f}" if m["axle_4_5"] is not None else "", normal),
            Paragraph(f"{float(m['axle_6']):.2f}" if m["axle_6"] is not None else "", normal),
            Paragraph(f"{float(m['semi']):.2f}" if m["semi"] is not None else "", normal),
            Paragraph(f"{float(m['hydvac']):.2f}" if m["hydvac"] is not None else "", normal),
            Paragraph(f"{float(m['dirt_in']):.2f}" if m["dirt_in"] is not None else "", normal),
            Paragraph("Active" if m["active"] else "Inactive", normal),
        ])

    table = Table(
        table_data,
        colWidths=[90, 45, 50, 38, 38, 38, 38, 38, 38, 45, 38, 38, 45],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ALIGN", (2, 1), (11, -1), "RIGHT"),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ALIGN", (12, 1), (12, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

def save_pdf(ticket_number, pdf_bytes):
    year = app_now().year
    if AZURE_STORAGE_CONNECTION_STRING:
        blob_name = f"{AZURE_TICKETS_BLOB_PREFIX}/{year}/{ticket_number}.pdf" if AZURE_TICKETS_BLOB_PREFIX else f"{year}/{ticket_number}.pdf"
        try:
            blob_url = upload_pdf_to_blob(blob_name, pdf_bytes)
            if blob_url:
                return blob_url
        except Exception as exc:
            app.logger.warning("Azure Blob upload failed for ticket %s: %s. Falling back to local storage.", ticket_number, exc)

    year_dir = PDF_DIR / str(year)
    year_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = year_dir / f"{ticket_number}.pdf"
    with open(pdf_path, "wb") as f:
        f.write(pdf_bytes)
    return str(pdf_path)


def print_pdf_file(pdf_path):
    if os.name != "nt":
        raise RuntimeError("Automatic printing is currently implemented for Windows only.")
    os.startfile(pdf_path, "print")


def is_active_column_boolean(db, table_name):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT data_type
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = %s
              AND column_name = 'active'
            """,
            (table_name,),
        )
        row = cursor.fetchone()
    return bool(row) and row["data_type"] == "boolean"


def validate_ticket_edit_password(admin_password):
    configured_password = os.getenv("TICKET_EDIT_PASSWORD", "").strip()
    if not configured_password:
        return False, "Ticket edit password is not configured on server."
    if not hmac.compare_digest(admin_password, configured_password):
        return False, "Password failed."
    return True, ""


def parse_optional_int(value):
    text = str(value or "").strip()
    if not text:
        return None
    return int(text)


def to_datetime_local_value(value):
    if value is None:
        return ""
    try:
        dt = value if isinstance(value, datetime) else datetime.fromisoformat(str(value).strip())
        if dt.tzinfo is not None:
            dt = dt.astimezone(APP_TZ).replace(tzinfo=None)
        return dt.strftime("%Y-%m-%dT%H:%M")
    except Exception:
        return ""


def validate_material_admin_password(admin_password):
    configured_password = os.getenv("MATERIAL_ADMIN_PASSWORD", "").strip()
    if not configured_password:
        return False, "Material admin password is not configured on server."
    if not hmac.compare_digest(admin_password, configured_password):
        return False, "Password failed."
    return True, ""


def parse_material_active_value(raw_value, use_boolean):
    value = str(raw_value or "").strip().upper()
    if value in {"1", "TRUE", "T", "YES", "Y", "ACTIVE", "A"}:
        return True if use_boolean else 1
    if value in {"0", "FALSE", "F", "NO", "N", "INACTIVE", "I"}:
        return False if use_boolean else 0
    raise ValueError("active must be 1/0, TRUE/FALSE, YES/NO, ACTIVE/INACTIVE")


def parse_materials_upload_rows(uploaded_file, filename):
    ext = Path(filename or "").suffix.lower()

    if ext == ".csv":
        text_stream = io.TextIOWrapper(uploaded_file.stream, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(text_stream)
        rows = []
        for idx, row in enumerate(reader, start=2):
            rows.append((idx, row))
        return rows

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel upload requires the openpyxl package.") from exc

        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []

        headers = [str(col).strip() if col is not None else "" for col in header_row]
        rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for header, value in zip(headers, values):
                row_dict[header] = "" if value is None else str(value)
            rows.append((row_number, row_dict))
        return rows

    raise RuntimeError("Unsupported file type. Please upload a .csv or .xlsx file.")


def upsert_job_cache_row(db, job_code, job_name, customer, tax_exempt, active, source_updated_at, refreshed_at):
    with db.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO jobs_cache (job_code, job_name, customer, tax_exempt, active, source_updated_at, refreshed_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(job_code) DO UPDATE SET
                job_name = excluded.job_name,
                customer = excluded.customer,
                tax_exempt = excluded.tax_exempt,
                active = excluded.active,
                source_updated_at = excluded.source_updated_at,
                refreshed_at = excluded.refreshed_at
            """,
            (job_code, job_name, customer, tax_exempt, active, source_updated_at, refreshed_at),
        )


def refresh_jobs_cache(db):
    app.logger.info("Refreshing jobs cache...")
    csv_file = resolve_jobs_csv_path()
    if csv_file is not None:

        now = app_now().isoformat(timespec="seconds")
        synced = 0

        with open(csv_file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise RuntimeError("CSV is missing a header row.")

            def first_present(*names):
                for name in names:
                    if name in reader.fieldnames:
                        return name
                return None

            job_code_col = first_present("job_code", "Job #", "Job#", "Job Number")
            job_name_col = first_present("job_name", "Job Name")
            customer_col = first_present("customer", "Customer Name")
            tax_exempt_col = first_present("tax_exempt", "Tax Exempt", "TaxExempt")
            active_col = first_present("active", "Job Status")
            source_updated_at_col = first_present("source_updated_at")

            missing = []
            if not job_code_col:
                missing.append("job_code or Job #")
            if not job_name_col:
                missing.append("job_name or Job Name")
            if missing:
                raise RuntimeError(f"CSV missing required columns: {', '.join(missing)}")

            def parse_active_value(raw_value):
                value = str(raw_value or "").strip().upper()
                if not value:
                    return 1
                if value in {"1", "A", "ACTIVE", "Y", "YES", "TRUE", "T"}:
                    return 1
                if value in {"0", "I", "C", "INACTIVE", "N", "NO", "FALSE", "F"}:
                    return 0
                try:
                    return 1 if int(value) == 1 else 0
                except ValueError:
                    return 1

            for row in reader:
                job_code = str(row.get(job_code_col) or "").strip()
                if not job_code:
                    continue

                job_name = str(row.get(job_name_col) or "").strip()
                customer = str(row.get(customer_col) or "").strip() if customer_col else ""
                tax_exempt = str(row.get(tax_exempt_col) or "").strip() if tax_exempt_col else ""
                active_raw = str(row.get(active_col) or "").strip() if active_col else ""
                active = parse_active_value(active_raw)
                source_updated_at = (
                    str(row.get(source_updated_at_col) or "").strip() if source_updated_at_col else ""
                ) or None

                upsert_job_cache_row(
                    db=db,
                    job_code=job_code,
                    job_name=job_name,
                    customer=customer,
                    tax_exempt=tax_exempt,
                    active=active,
                    source_updated_at=source_updated_at,
                    refreshed_at=now,
                )
                synced += 1
        app.logger.info("Jobs cache refreshed from CSV. %s rows synced.", synced)

        return synced

    odbc_conn = os.getenv("REMOTE_SQL_ODBC_CONNECTION_STRING")
    if not odbc_conn:
        raise RuntimeError(
            "No jobs source found. Set JOBS_CSV_PATH, place CSV at data/jobs.csv, "
            "or configure REMOTE_SQL_ODBC_CONNECTION_STRING."
        )

    query = os.getenv(
        "JOBS_SQL_QUERY",
        """
        SELECT
            CAST(job_code AS NVARCHAR(100)) AS job_code,
            CAST(job_name AS NVARCHAR(255)) AS job_name,
            CAST(customer AS NVARCHAR(255)) AS customer,
            CAST(active AS INT) AS active,
            source_updated_at
        FROM jobs
        """,
    )

    with pyodbc.connect(odbc_conn, timeout=10) as conn:
        cursor = conn.cursor()
        rows = cursor.execute(query).fetchall()

    now = app_now().isoformat(timespec="seconds")
    for row in rows:
        job_code = str(row.job_code).strip()
        job_name = str(row.job_name).strip() if row.job_name is not None else ""
        customer = str(row.customer).strip() if row.customer is not None else ""
        active = int(row.active) if row.active is not None else 1
        source_updated_at = (
            str(row.source_updated_at) if getattr(row, "source_updated_at", None) is not None else None
        )

        upsert_job_cache_row(
            db=db,
            job_code=job_code,
            job_name=job_name,
            customer=customer,
            tax_exempt="",
            active=active,
            source_updated_at=source_updated_at,
            refreshed_at=now,
        )
    return len(rows)

def list_jobs(db):
    with db.cursor() as cursor:
        cursor.execute(
        """
        SELECT id, job_code, job_name, customer
        FROM jobs_cache
        ORDER BY job_code
        """
        )
        return cursor.fetchall()


def list_ticket_jobs(db):
    with db.cursor() as cursor:
        cursor.execute(
        """
        WITH manual AS (
            SELECT
                ('manual:' || id::text) AS job_key,
                job_code,
                job_name,
                customer,
                tax_exempt
            FROM manual_jobs
        ),
        cache_all AS (
            SELECT
                ('cache:' || c.id::text) AS job_key,
                c.job_code,
                c.job_name,
                c.customer,
                c.tax_exempt
            FROM jobs_cache c
            WHERE NOT EXISTS (
                SELECT 1
                FROM manual m
                WHERE m.job_code = c.job_code
                  AND m.job_name = c.job_name
            )
        )
        SELECT job_key, job_code, job_name, customer, tax_exempt
        FROM manual

        UNION ALL

        SELECT job_key, job_code, job_name, customer, tax_exempt
        FROM cache_all

        ORDER BY job_code, job_name
        """
        )
        return cursor.fetchall()


def search_ticket_jobs(db, query_text=""):
    search_term = str(query_text or "").strip()
    like_pattern = f"%{search_term}%"

    with db.cursor() as cursor:
        cursor.execute(
            """
            WITH manual AS (
                SELECT
                    ('manual:' || id::text) AS job_key,
                    COALESCE(job_code, '') AS job_code,
                    COALESCE(job_name, '') AS job_name,
                    COALESCE(customer, '') AS customer,
                    COALESCE(tax_exempt, '') AS tax_exempt
                FROM manual_jobs
            ),
            cache_all AS (
                SELECT
                    ('cache:' || c.id::text) AS job_key,
                    COALESCE(c.job_code, '') AS job_code,
                    COALESCE(c.job_name, '') AS job_name,
                    COALESCE(c.customer, '') AS customer,
                    COALESCE(c.tax_exempt, '') AS tax_exempt
                FROM jobs_cache c
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM manual m
                    WHERE m.job_code = c.job_code
                      AND m.job_name = c.job_name
                )
            ),
            merged AS (
                SELECT * FROM manual
                UNION ALL
                SELECT * FROM cache_all
            )
            SELECT job_key, job_code, job_name, customer, tax_exempt
            FROM merged
            WHERE (
                %s = ''
                OR job_code ILIKE %s
                OR job_name ILIKE %s
                OR customer ILIKE %s
                OR (job_code || ' - ' || job_name) ILIKE %s
            )
            ORDER BY job_code, job_name
            """,
            (search_term, like_pattern, like_pattern, like_pattern, like_pattern),
        )
        return cursor.fetchall()


def list_recent_tickets(db, limit=5):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                id,
                ticket_number,
                created_at,
                direction,
                customer_snapshot,
                truck_number_snapshot,
                material_name_snapshot,
                quantity,
                unit
            FROM tickets
            WHERE COALESCE(active, TRUE) = TRUE
            ORDER BY created_at DESC, id DESC
            LIMIT %s
            """,
            (limit,),
        )
        return cursor.fetchall()


def split_job_entry(job_entry):
    if " - " in job_entry:
        job_code, job_name = [part.strip() for part in job_entry.split(" - ", 1)]
    else:
        job_code = job_entry.strip()
        job_name = job_entry.strip()
    return job_code, job_name


def get_or_create_manual_job(db, job_entry):
    job_code, job_name = split_job_entry(job_entry)
    now = app_now().isoformat(timespec="seconds")
    with db.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO manual_jobs (job_code, job_name, customer, tax_exempt, active, created_at)
            VALUES (%s, %s, %s, %s, 1, %s)
            ON CONFLICT (job_code, job_name)
            DO UPDATE SET active = 1
            RETURNING id, job_code, job_name, customer, tax_exempt
            """,
            (job_code, job_name, "", "New", now),
        )
        return cursor.fetchone()


def get_selected_job(db, selected_job_id):
    if not selected_job_id:
        return None, None

    if selected_job_id.startswith("cache:"):
        job_id = selected_job_id.split(":", 1)[1]
        if not job_id.isdigit():
            return None, None
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, job_code, job_name, customer, tax_exempt FROM jobs_cache WHERE id = %s",
                (job_id,),
            )
            return cursor.fetchone(), "cache"

    if selected_job_id.startswith("manual:"):
        job_id = selected_job_id.split(":", 1)[1]
        if not job_id.isdigit():
            return None, None
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, job_code, job_name, customer, tax_exempt FROM manual_jobs WHERE id = %s",
                (job_id,),
            )
            return cursor.fetchone(), "manual"

    if selected_job_id.isdigit():
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, job_code, job_name, customer, tax_exempt FROM jobs_cache WHERE id = %s",
                (selected_job_id,),
            )
            return cursor.fetchone(), "cache"

    return None, None


def apply_reports_job_filter(db, where, params, selected_job_id="", job_entry=""):
    selected_job_id = str(selected_job_id or "").strip()
    job_entry = str(job_entry or "").strip()

    if selected_job_id:
        selected_job, selected_source = get_selected_job(db, selected_job_id)
        if selected_job:
            if selected_source == "cache":
                where.append("t.job_id = %s")
                params.append(selected_job["id"])
                return

            where.append("t.job_id IS NULL")
            where.append("COALESCE(t.job_code_snapshot, '') = %s")
            params.append(str(selected_job.get("job_code") or "").strip())
            where.append("COALESCE(t.job_name_snapshot, '') = %s")
            params.append(str(selected_job.get("job_name") or "").strip())
            return

        if selected_job_id.isdigit():
            where.append("t.job_id = %s")
            params.append(selected_job_id)
            return

    if job_entry:
        like_term = f"%{job_entry}%"
        where.append(
            "(" 
            "COALESCE(t.job_code_snapshot, '') ILIKE %s OR "
            "COALESCE(t.job_name_snapshot, '') ILIKE %s OR "
            "(COALESCE(t.job_code_snapshot, '') || ' - ' || COALESCE(t.job_name_snapshot, '')) ILIKE %s"
            ")"
        )
        params.extend([like_term, like_term, like_term])


def get_job_entry_display(db, selected_job_id):
    selected_job, _selected_source = get_selected_job(db, selected_job_id)
    if not selected_job:
        return ""
    job_code = str(selected_job.get("job_code") or "").strip()
    job_name = str(selected_job.get("job_name") or "").strip()
    if job_code and job_name:
        return f"{job_code} - {job_name}"
    return job_code or job_name


def apply_reports_customer_filter(db, where, params, customer_id="", customer=""):
    customer_id = str(customer_id or "").strip()
    customer = str(customer or "").strip()

    if customer_id.isdigit():
        with db.cursor() as cursor:
            cursor.execute("SELECT customer_name FROM customers WHERE id = %s", (customer_id,))
            customer_row = cursor.fetchone()
        if customer_row and str(customer_row.get("customer_name") or "").strip():
            where.append("LOWER(TRIM(COALESCE(t.customer_snapshot, ''))) = LOWER(TRIM(%s))")
            params.append(str(customer_row.get("customer_name") or "").strip())
            return

    if customer:
        where.append("t.customer_snapshot ILIKE %s")
        params.append(f"%{customer}%")


def get_customer_display_by_id(db, customer_id):
    customer_id = str(customer_id or "").strip()
    if not customer_id.isdigit():
        return ""
    with db.cursor() as cursor:
        cursor.execute("SELECT customer_name FROM customers WHERE id = %s", (customer_id,))
        customer_row = cursor.fetchone()
    if not customer_row:
        return ""
    return str(customer_row.get("customer_name") or "").strip()


def apply_reports_material_filter(where, params, material_id="", material_entry=""):
    material_id = str(material_id or "").strip()
    material_entry = str(material_entry or "").strip()

    if material_id.isdigit():
        where.append("t.material_id = %s")
        params.append(material_id)
        return

    if material_entry:
        where.append("COALESCE(t.material_name_snapshot, '') ILIKE %s")
        params.append(f"%{material_entry}%")


def get_material_entry_display(db, material_id):
    material_id = str(material_id or "").strip()
    if not material_id.isdigit():
        return ""
    with db.cursor() as cursor:
        cursor.execute("SELECT material FROM material_price WHERE id = %s", (material_id,))
        material_row = cursor.fetchone()
    if not material_row:
        return ""
    return str(material_row.get("material") or "").strip()

# def list_trucks(db):
#     return db.execute(
#         "SELECT id, truck_number, description, truck_size, hauled_by, active FROM trucks WHERE active = 1 ORDER BY truck_number"
#     ).fetchall()
def list_customers(db):
    with db.cursor() as cursor:
        cursor.execute(
        "SELECT id, customer_name FROM customers ORDER BY customer_name"
        )
        return cursor.fetchall()


def get_customer_by_name(db, customer_name):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, customer_name
            FROM customers
            WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (customer_name,),
        )
        return cursor.fetchone()


def list_trucks(db):
    with db.cursor() as cursor:
        cursor.execute(
        "SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, active FROM trucks_main WHERE active = TRUE ORDER BY truck_number"
        )
        return cursor.fetchall()


def list_materials(db, direction=None):
    with db.cursor() as cursor:
        if direction:
            cursor.execute(
                "SELECT id, material AS material_name, active, direction FROM material_price WHERE active = TRUE AND direction = %s ORDER BY material_name",
                (direction,),
            )
        else:
            cursor.execute(
                "SELECT id, material AS material_name, active, direction FROM material_price WHERE active = TRUE ORDER BY material_name"
            )
        return cursor.fetchall()


@app.before_request
def require_login():
    if request.endpoint is None:
        return

    open_endpoints = {
        "login",
        "logout",
        "static",
        "healthz",
        "api_credit_card_daily_report",
        "api_non_credit_card_daily_report",
        "api_notification_truck_seen",
        "test"
    }
    if request.endpoint in open_endpoints:
        return
    if not session.get("logged_in"):
        return redirect(url_for("login", next=request.url))


def parse_iso_datetime(value):
    text = str(value or "").strip()
    if not text:
        return None
    if re.fullmatch(r"[0-9]{13}", text):
        return datetime.fromtimestamp(int(text) / 1000.0, tz=timezone.utc)
    if re.fullmatch(r"[0-9]{10}(?:\.[0-9]+)?", text):
        return datetime.fromtimestamp(float(text), tz=timezone.utc)
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    return datetime.fromisoformat(text)


def parse_webhook_payload():
    body = request.get_json(silent=True)
    if body is not None:
        return body

    raw_body = request.get_data(as_text=True) or ""
    if raw_body.strip():
        try:
            return json.loads(raw_body)
        except Exception:
            pass

    if request.form:
        return request.form.to_dict(flat=True)
    return {}


def payload_first_value(payload, keys):
    normalized_keys = {str(k).strip().lower().replace("-", "_") for k in keys}
    stack = [payload]

    while stack:
        current = stack.pop()
        if isinstance(current, dict):
            for key, value in current.items():
                key_norm = str(key).strip().lower().replace("-", "_")
                if key_norm in normalized_keys and value is not None:
                    if isinstance(value, (str, int, float, bool)):
                        text = str(value).strip()
                        if text:
                            return text
                if isinstance(value, (dict, list, tuple)):
                    stack.append(value)
        elif isinstance(current, (list, tuple)):
            stack.extend(current)

    return ""


# def normalize_epc(value):
#     text = str(value or "").strip()
#     if not text:
#         return ""
#     if text.lower().startswith("0x"):
#         text = text[2:]
#     return text.replace(" ", "").replace("-", "").upper()

def normalize_epc(value):
    text = str(value or "").strip()
    if not text:
        return ""

    if text.lower().startswith("0x"):
        text = text[2:]

    text_no_sep = text.replace(" ", "").replace("-", "")

    if re.fullmatch(r"[0-9A-Fa-f]+", text_no_sep):
        return text_no_sep.upper()

    try:
        decoded = base64.b64decode(text_no_sep, validate=True)
        if decoded:
            return decoded.hex().upper()
    except Exception:
        pass

    return text_no_sep.upper()

def truck_number_from_epc(db, epc_raw):
    epc = normalize_epc(epc_raw)
    if not epc:
        return ""

    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT truck_number
            FROM rfid_epc_truck_map
            WHERE COALESCE(active, TRUE) = TRUE
              AND LOWER(TRIM(epc)) = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (epc,),
        )
        row = cursor.fetchone()
    if not row:
        return ""
    return str(row.get("truck_number") or "").strip()


def normalize_serial_number(value):
    text = str(value or "").strip()
    if re.fullmatch(r"[0-9]+\.0", text):
        text = text[:-2]
    return text.upper()


def resolve_rfid_upload_columns(sample_row):
    key_lookup = {str(k).strip().lower(): k for k in sample_row.keys()}
    epc_col = key_lookup.get("epc") or key_lookup.get("epc_hex") or key_lookup.get("epchex")
    serial_col = (
        key_lookup.get("serial_number")
        or key_lookup.get("serial")
        or key_lookup.get("sticker_number")
        or key_lookup.get("sticker")
    )
    return epc_col, serial_col


def upsert_rfid_tag_catalog_row(db, serial_number, epc):
    with db.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO rfid_tag_catalog (serial_number, epc, active, updated_at)
            VALUES (%s, %s, TRUE, NOW())
            ON CONFLICT (serial_number) DO UPDATE
            SET epc = EXCLUDED.epc,
                active = TRUE,
                updated_at = NOW()
            """,
            (serial_number, epc),
        )


def list_rfid_assignments(db):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                a.id,
                a.serial_number,
                c.epc,
                a.truck_id,
                t.truck_number,
                a.notes,
                a.assigned_at,
                a.updated_at
            FROM rfid_truck_tag_assignment a
            JOIN rfid_tag_catalog c ON c.serial_number = a.serial_number
            JOIN trucks_main t ON t.id = a.truck_id
            ORDER BY t.truck_number
            """
        )
        return cursor.fetchall()


def list_unassigned_rfid_tags(db):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT c.serial_number, c.epc
            FROM rfid_tag_catalog c
            LEFT JOIN rfid_truck_tag_assignment a ON a.serial_number = c.serial_number
            WHERE COALESCE(c.active, TRUE) = TRUE
              AND a.id IS NULL
            ORDER BY c.serial_number
            """
        )
        return cursor.fetchall()


def list_unassigned_trucks(db):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT t.id, t.truck_number
            FROM trucks_main t
            LEFT JOIN rfid_truck_tag_assignment a ON a.truck_id = t.id
            WHERE COALESCE(t.active, TRUE) = TRUE
              AND a.id IS NULL
            ORDER BY t.truck_number
            """
        )
        return cursor.fetchall()


def sync_epc_map_for_assignment(db, serial_number):
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT c.epc, t.truck_number
            FROM rfid_truck_tag_assignment a
            JOIN rfid_tag_catalog c ON c.serial_number = a.serial_number
            JOIN trucks_main t ON t.id = a.truck_id
            WHERE a.serial_number = %s
            LIMIT 1
            """,
            (serial_number,),
        )
        row = cursor.fetchone()

    if not row:
        return

    epc = normalize_epc(row.get("epc"))
    truck_number = str(row.get("truck_number") or "").strip()
    if not epc or not truck_number:
        return

    with db.cursor() as cursor:
        cursor.execute(
            """
            UPDATE rfid_epc_truck_map
            SET active = FALSE,
                updated_at = NOW()
            WHERE LOWER(TRIM(truck_number)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(epc)) <> LOWER(TRIM(%s))
            """,
            (truck_number, epc),
        )
        cursor.execute(
            """
            INSERT INTO rfid_epc_truck_map (epc, truck_number, source, active, updated_at)
            VALUES (%s, %s, %s, TRUE, NOW())
            ON CONFLICT (epc)
            DO UPDATE SET
                truck_number = EXCLUDED.truck_number,
                source = EXCLUDED.source,
                active = TRUE,
                updated_at = NOW()
            """,
            (epc, truck_number, "serial_assignment"),
        )


def deactivate_epc_map_for_serial(db, serial_number):
    with db.cursor() as cursor:
        cursor.execute("SELECT epc FROM rfid_tag_catalog WHERE serial_number = %s LIMIT 1", (serial_number,))
        row = cursor.fetchone()
    if not row:
        return

    epc = normalize_epc(row.get("epc"))
    if not epc:
        return

    with db.cursor() as cursor:
        cursor.execute(
            """
            UPDATE rfid_epc_truck_map
            SET active = FALSE,
                updated_at = NOW()
            WHERE LOWER(TRIM(epc)) = LOWER(TRIM(%s))
            """,
            (epc,),
        )


def search_rfid_assignments(db, query):
    query_text = str(query or "").strip()
    with db.cursor() as cursor:
        if not query_text:
            cursor.execute(
                """
                SELECT
                    a.id,
                    a.serial_number,
                    c.epc,
                    t.truck_number,
                    a.notes,
                    a.assigned_at,
                    a.updated_at
                FROM rfid_truck_tag_assignment a
                JOIN rfid_tag_catalog c ON c.serial_number = a.serial_number
                JOIN trucks_main t ON t.id = a.truck_id
                ORDER BY a.updated_at DESC
                LIMIT 100
                """
            )
            return cursor.fetchall()

        wildcard = f"%{query_text}%"
        cursor.execute(
            """
            SELECT
                a.id,
                a.serial_number,
                c.epc,
                t.truck_number,
                a.notes,
                a.assigned_at,
                a.updated_at
            FROM rfid_truck_tag_assignment a
            JOIN rfid_tag_catalog c ON c.serial_number = a.serial_number
            JOIN trucks_main t ON t.id = a.truck_id
            WHERE t.truck_number ILIKE %s
               OR a.serial_number ILIKE %s
               OR c.epc ILIKE %s
            ORDER BY a.updated_at DESC
            LIMIT 200
            """,
            (wildcard, wildcard, wildcard),
        )
        return cursor.fetchall()


def notification_in_cooldown(db, truck_number, source, cooldown_seconds):
    if cooldown_seconds <= 0:
        return False

    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id
            FROM rfid_notifications
            WHERE event_type = 'known_truck_detected'
              AND LOWER(TRIM(truck_number)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(source)) = LOWER(TRIM(%s))
              AND created_at >= NOW() - (%s * INTERVAL '1 second')
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (truck_number, source, cooldown_seconds),
        )
        return bool(cursor.fetchone())

@app.route("/api/rfid/root-test", methods=["GET", "POST"])
def rfid_root_test():
    print("R700 reached root")
    print(request.method)
    print(dict(request.headers))
    print(request.get_data(as_text=True))
    return {"ok": True}, 200

@app.post("/api/notifications/truck-seen/test")
def test():
    print("=" * 80)
    print("Headers:", dict(request.headers))
    print("Raw:", request.get_data())
    print("Text:", request.get_data(as_text=True))
    print("JSON:", request.get_json(silent=True))
    return "", 200

@app.post("/api/notifications/truck-seen")
def api_notification_truck_seen():

    if not NOTIFICATIONS_ENABLED:
        return {"ok": False, "error": "Notifications are temporarily disabled."}, 503

    provided_key = (request.headers.get("X-API-Key") or request.args.get("api_key") or "").strip()
    auth = request.authorization
    basic_auth_configured = bool(RFID_WEBHOOK_USERNAME or RFID_WEBHOOK_PASSWORD)
    basic_auth_ok = bool(
        auth
        and hmac.compare_digest(str(auth.username or ""), RFID_WEBHOOK_USERNAME)
        and hmac.compare_digest(str(auth.password or ""), RFID_WEBHOOK_PASSWORD)
    )

    if RFID_EVENT_API_KEY:
        key_ok = hmac.compare_digest(provided_key, RFID_EVENT_API_KEY)
        if not (key_ok or (basic_auth_configured and basic_auth_ok)):
            return {"ok": False, "error": "Unauthorized. Provide valid API key or webhook basic auth."}, 401
    elif basic_auth_configured:
        if not basic_auth_ok:
            return {"ok": False, "error": "Unauthorized. Invalid webhook basic auth."}, 401
    else:
        forwarded_for = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
        remote_ip = forwarded_for or (request.remote_addr or "")
        if remote_ip not in {"127.0.0.1", "::1", "localhost"}:
            return {
                "ok": False,
                "error": "Configure RFID_EVENT_API_KEY or webhook basic auth, or call from localhost.",
            }, 401

    body = parse_webhook_payload()
    events = body if isinstance(body, list) else [body or {}]

    fallback_source = (
        payload_first_value(body, ["source", "reader_name", "reader", "device_name", "hostname"])
        or request.args.get("source", "").strip()
        or "r700-gate-1"
    )
    fallback_message = payload_first_value(body, ["message", "note", "description"]) or str(request.args.get("message") or "").strip()
    fallback_truck_number = payload_first_value(body, ["truck_number", "truck", "truck_no", "truckno"]) or str(request.args.get("truck_number") or "").strip()
    fallback_detected_at_raw = payload_first_value(body, ["detected_at", "timestamp", "event_time", "time", "first_seen_time"]) or str(request.args.get("detected_at") or "").strip()
    fallback_epc_raw = payload_first_value(body, ["epc", "tag_epc", "tagEpc", "epc_hex", "id_hex", "tag_id", "tagid"]) or request.args.get("epc", "").strip()

    db = get_db()
    truck_cache = {}
    seen_payload_keys = set()
    created_notifications = []
    suppressed = []
    skipped = []
    mapped_any = False

    for index, event in enumerate(events):
        if not isinstance(event, dict):
            skipped.append({"index": index, "reason": "event_not_object"})
            continue

        tag_event = event.get("tagInventoryEvent", {}) if isinstance(event.get("tagInventoryEvent"), dict) else {}
        source = (
            payload_first_value(event, ["source", "reader_name", "reader", "device_name", "hostname"])
            or fallback_source
        )
        source = str(source or "r700-gate-1").strip() or "r700-gate-1"

        custom_message = payload_first_value(event, ["message", "note", "description"]) or fallback_message
        truck_number = payload_first_value(event, ["truck_number", "truck", "truck_no", "truckno"]) or fallback_truck_number
        detected_at_raw = (
            event.get("timestamp")
            or payload_first_value(event, ["detected_at", "timestamp", "event_time", "time", "first_seen_time"])
            or fallback_detected_at_raw
        )
        epc_raw = (
            tag_event.get("epcHex")
            or tag_event.get("epc")
            or payload_first_value(event, ["epc", "tag_epc", "tagEpc", "epc_hex", "id_hex", "tag_id", "tagid"])
            or fallback_epc_raw
        )

        epc_normalized = normalize_epc(epc_raw)
        payload_key = (f"epc:{epc_normalized}" if epc_normalized else f"truck:{str(truck_number).strip().lower()}")
        if payload_key in seen_payload_keys:
            skipped.append({"index": index, "reason": "duplicate_in_payload", "epc": epc_normalized or None})
            continue
        seen_payload_keys.add(payload_key)

        mapped_from_epc = False
        truck_number = str(truck_number or "").strip()
        if not truck_number and epc_normalized:
            truck_number = truck_number_from_epc(db, epc_normalized)
            mapped_from_epc = bool(truck_number)
            mapped_any = mapped_any or mapped_from_epc

        print("event_index:", index)
        print("epc_raw:", epc_raw)
        print("epc_normalized:", epc_normalized)
        print("truck_number before mapping:", truck_number)

        if not truck_number:
            skipped.append(
                {
                    "index": index,
                    "reason": "no_truck_mapping",
                    "epc": epc_normalized or None,
                }
            )
            continue

        truck_cache_key = truck_number.strip().lower()
        truck = truck_cache.get(truck_cache_key)
        if truck is None:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT id, truck_number
                    FROM trucks_main
                    WHERE active = TRUE AND LOWER(TRIM(truck_number)) = LOWER(TRIM(%s))
                    LIMIT 1
                    """,
                    (truck_number,),
                )
                truck = cursor.fetchone()
            truck_cache[truck_cache_key] = truck

        if not truck:
            skipped.append({"index": index, "reason": "truck_not_active", "truck_number": truck_number})
            continue

        normalized_truck_number = str(truck.get("truck_number") or truck_number).strip() or truck_number
        detected_at = None
        if detected_at_raw:
            try:
                detected_at = parse_iso_datetime(detected_at_raw)
            except ValueError:
                skipped.append({"index": index, "reason": "invalid_detected_at", "value": str(detected_at_raw)})
                continue

        if notification_in_cooldown(db, normalized_truck_number, source, RFID_NOTIFICATION_COOLDOWN_SECONDS):
            suppressed.append(
                {
                    "index": index,
                    "truck_number": normalized_truck_number,
                    "epc": epc_normalized or None,
                    "cooldown_seconds": RFID_NOTIFICATION_COOLDOWN_SECONDS,
                }
            )
            continue

        message = custom_message or f"Known truck appeared: {normalized_truck_number}"
        if epc_normalized:
            message = custom_message or f"Known truck appeared: {normalized_truck_number} (EPC: {epc_normalized})"

        try:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO rfid_notifications (event_type, truck_number, source, message, detected_at)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id, status, created_at
                    """,
                    ("known_truck_detected", normalized_truck_number, source, message, detected_at),
                )
                created = cursor.fetchone()
        except Exception as exc:
            db.rollback()
            return {"ok": False, "error": f"Could not store notification: {exc}"}, 500

        created_at = created.get("created_at")
        created_item = {
            "id": int(created.get("id") or 0),
            "event_type": "known_truck_detected",
            "truck_number": normalized_truck_number,
            "source": source,
            "message": message,
            "status": created.get("status") or "pending",
            "detected_at": detected_at.isoformat() if detected_at else None,
            "created_at": created_at.isoformat() if created_at else None,
            "epc": epc_normalized or None,
            "mapped_from_epc": mapped_from_epc,
        }
        created_notifications.append(created_item)

    if created_notifications:
        db.commit()

    if not created_notifications and suppressed:
        return {
            "ok": True,
            "created_count": 0,
            "suppressed_count": len(suppressed),
            "skipped_count": len(skipped),
            "cooldown_seconds": RFID_NOTIFICATION_COOLDOWN_SECONDS,
            "suppressed": suppressed,
            "skipped": skipped,
        }, 200

    if not created_notifications:
        epc_hint = next((item.get("epc") for item in skipped if item.get("epc")), None)
        return {
            "ok": False,
            "error": "No valid mapped truck events were found in this payload.",
            "epc": epc_hint,
            "skipped": skipped,
        }, 404

    response_payload = {
        "ok": True,
        "mapped_from_epc": mapped_any,
        "created_count": len(created_notifications),
        "suppressed_count": len(suppressed),
        "skipped_count": len(skipped),
        "cooldown_seconds": RFID_NOTIFICATION_COOLDOWN_SECONDS,
        "notifications": created_notifications,
        "suppressed": suppressed,
        "skipped": skipped,
    }

    if len(created_notifications) == 1:
        response_payload["notification"] = created_notifications[0]
        response_payload["epc"] = created_notifications[0].get("epc")

    print("RFID Notification Created:")
    print(response_payload)

    return response_payload, 201


def serialize_notification_row(row):
    detected_at = row.get("detected_at")
    created_at = row.get("created_at")
    decided_at = row.get("decided_at")
    return {
        "id": int(row.get("id") or 0),
        "event_type": row.get("event_type") or "known_truck_detected",
        "truck_number": row.get("truck_number") or "",
        "source": row.get("source") or "rfid",
        "message": row.get("message") or "",
        "status": row.get("status") or "pending",
        "detected_at": detected_at.isoformat() if detected_at else None,
        "created_at": created_at.isoformat() if created_at else None,
        "decided_at": decided_at.isoformat() if decided_at else None,
    }


@app.get("/api/notifications/poll")
def api_notifications_poll():
    if not NOTIFICATIONS_ENABLED:
        return {"ok": False, "error": "Notifications are temporarily disabled."}, 503

    since_id_raw = str(request.args.get("since_id") or "0").strip()
    limit_raw = str(request.args.get("limit") or "20").strip()

    try:
        since_id = max(0, int(since_id_raw))
    except ValueError:
        return {"ok": False, "error": "Invalid since_id."}, 400

    try:
        limit = int(limit_raw)
    except ValueError:
        return {"ok": False, "error": "Invalid limit."}, 400

    limit = max(1, min(limit, 100))

    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, event_type, truck_number, source, message, status, detected_at, created_at, decided_at
            FROM rfid_notifications
            WHERE id > %s
            ORDER BY id ASC
            LIMIT %s
            """,
            (since_id, limit),
        )
        rows = cursor.fetchall()

    notifications = []
    latest_id = since_id
    for row in rows:
        row_id = int(row.get("id") or 0)
        latest_id = max(latest_id, row_id)
        notifications.append(serialize_notification_row(row))

    return {"ok": True, "notifications": notifications, "latest_id": latest_id}, 200


@app.get("/api/notifications/list")
def api_notifications_list():
    if not NOTIFICATIONS_ENABLED:
        return {"ok": False, "error": "Notifications are temporarily disabled."}, 503

    status = str(request.args.get("status") or "all").strip().lower()
    limit_raw = str(request.args.get("limit") or "100").strip()

    try:
        limit = int(limit_raw)
    except ValueError:
        return {"ok": False, "error": "Invalid limit."}, 400

    limit = max(1, min(limit, 300))

    where_sql = ""
    params = []
    if status in {"pending", "approved", "denied"}:
        where_sql = "WHERE status = %s"
        params.append(status)
    elif status != "all":
        return {"ok": False, "error": "Invalid status. Use all, pending, approved, or denied."}, 400

    query = (
        "SELECT id, event_type, truck_number, source, message, status, detected_at, created_at, decided_at "
        f"FROM rfid_notifications {where_sql} "
        "ORDER BY id DESC LIMIT %s"
    )
    params.append(limit)

    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()

    notifications = [serialize_notification_row(row) for row in rows]

    with db.cursor() as cursor:
        cursor.execute("SELECT COUNT(1) AS pending_count FROM rfid_notifications WHERE status = 'pending'")
        count_row = cursor.fetchone() or {}
    pending_count = int(count_row.get("pending_count") or 0)

    return {"ok": True, "notifications": notifications, "pending_count": pending_count}, 200


@app.post("/api/notifications/<int:notification_id>/decision")
def api_notifications_decision(notification_id):
    if not NOTIFICATIONS_ENABLED:
        return {"ok": False, "error": "Notifications are temporarily disabled."}, 503

    body = request.get_json(silent=True) if request.is_json else {}
    body = body or {}
    decision = str(body.get("decision") or request.form.get("decision") or "").strip().lower()
    if decision not in {"approve", "deny", "approved", "denied"}:
        return {"ok": False, "error": "decision must be approve or deny."}, 400

    target_status = "approved" if decision in {"approve", "approved"} else "denied"
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            UPDATE rfid_notifications
            SET status = %s,
                decided_at = NOW()
            WHERE id = %s
              AND status = 'pending'
            RETURNING id, event_type, truck_number, source, message, status, detected_at, created_at, decided_at
            """,
            (target_status, notification_id),
        )
        row = cursor.fetchone()

    if not row:
        db.rollback()
        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, event_type, truck_number, source, message, status, detected_at, created_at, decided_at
                FROM rfid_notifications
                WHERE id = %s
                LIMIT 1
                """,
                (notification_id,),
            )
            existing = cursor.fetchone()
        if not existing:
            return {"ok": False, "error": "Notification not found."}, 404
        return {
            "ok": False,
            "error": f"Notification already {existing.get('status') or 'processed'}.",
            "notification": serialize_notification_row(existing),
        }, 409

    db.commit()
    return {"ok": True, "notification": serialize_notification_row(row)}, 200


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect(url_for("new_ticket"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if (
            APP_USERNAME
            and APP_PASSWORD
            and hmac.compare_digest(username, APP_USERNAME)
            and hmac.compare_digest(password, APP_PASSWORD)
        ):
            session["logged_in"] = True
            session.permanent = True
            next_url = request.args.get("next") or url_for("new_ticket")
            parsed_next = urlparse(next_url)
            if parsed_next.netloc:
                next_url = url_for("new_ticket")
            return redirect(next_url)
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.post("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def home():
    return redirect(url_for("new_ticket"))


@app.get("/healthz")
def healthz():
    return "ok", 200


@app.route("/tickets/new", methods=["GET", "POST"])
def new_ticket():
    db = get_db()
    created_ticket_id = None

    if request.method == "POST":
        direction = request.form.get("direction", "IN").strip().upper()
        job_id = request.form.get("job_id", "").strip()
        job_entry = request.form.get("job_entry", "").strip()
        truck_id = request.form.get("truck_id", "").strip()
        truck_entry = request.form.get("truck_entry", "").strip()
        material_id = request.form.get("material_id", "").strip()
        material_entry = request.form.get("material_entry", "").strip()
        customer_id = request.form.get("customer_id", "").strip()
        quantity = (request.form.get("quantity") or "1").strip()
        unit = (request.form.get("unit") or "Load").strip()
        notes = request.form.get("notes", "").strip()
        auto_print = request.form.get("auto_print") == "on"
        use_now = request.form.get("use_now") == "on"
        custom_datetime = request.form.get("custom_datetime")

        if direction not in {"IN", "OUT"}:
            flash("Direction must be IN or OUT.", "error")
            return redirect(url_for("new_ticket"))
        if not all([job_entry, truck_entry, material_entry, quantity, unit]):
            flash("Job, truck, material, quantity, and unit are required.", "error")
            return redirect(url_for("new_ticket"))
        if not all([truck_id, material_id, customer_id]):
            flash("Please select customer, truck, and material from dropdown lists.", "error")
            return redirect(url_for("new_ticket"))
        # print(f"Received new ticket data: direction={direction}, job_id={job_id}, job_entry={job_entry}, truck_id={truck_id}, truck_entry={truck_entry}, material_id={material_id}, material_entry={material_entry}, customer_id={customer_id}, quantity={quantity}, unit={unit}, notes={notes}, auto_print={auto_print}, use_now={use_now}, custom_datetime={custom_datetime}")

        job = None
        truck = None
        material = None
        customer = None
        customer_snapshot = ""
        tax_exempt_snapshot = ""
        selected_job_source = None
        if job_id:
            job, selected_job_source = get_selected_job(db, job_id)
        if truck_id:
            with db.cursor() as cursor:
                cursor.execute("SELECT id, truck_number, truck_size FROM trucks_main WHERE id = %s", (truck_id,))
                truck = cursor.fetchone()
        if material_id:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT id, material AS material_name,
                           axle_1, tandem, triaxle, axle_4_5, axle_6, semi, hydvac, dirt_in
                    FROM material_price
                    WHERE id = %s
                    """,
                    (material_id,),
                )
                material = cursor.fetchone()
        if customer_id:
            with db.cursor() as cursor:
                cursor.execute("SELECT id, customer_name FROM customers WHERE id = %s", (customer_id,))
                customer = cursor.fetchone()
            if customer:
                customer_snapshot = (customer["customer_name"] or "").strip()

        if job:
            job_id_value = job["id"] if selected_job_source == "cache" else None
            job_code_snapshot = job["job_code"]
            job_name_snapshot = job["job_name"]
            tax_exempt_snapshot = (job["tax_exempt"] or "").strip()
            if not customer_snapshot:
                customer_snapshot = (job["customer"] or "").strip()
        else:
            manual_job = get_or_create_manual_job(db, job_entry)
            job_id_value = None
            job_code_snapshot = manual_job["job_code"]
            job_name_snapshot = manual_job["job_name"]
            tax_exempt_snapshot = (manual_job["tax_exempt"] or "").strip() or "New"
            if not customer_snapshot:
                customer_snapshot = (manual_job["customer"] or "").strip()

        if truck:
            truck_id_value = truck["id"]
            truck_number_snapshot = truck["truck_number"]
        else:
            truck_id_value = None
            truck_number_snapshot = truck_entry

        if material:
            material_id_value = material["id"]
            material_name_snapshot = material["material_name"]
        else:
            material_id_value = None
            material_name_snapshot = material_entry

        try:
            quantity_num = float(quantity)
        except ValueError:
            flash("Quantity must be numeric.", "error")
            return redirect(url_for("new_ticket"))

        ticket_cost = calculate_ticket_cost(truck, material, quantity_num)

        try:
            ticket_number, ticket_year, seq = next_ticket_number(db)
            if use_now:
                created_at = app_now().isoformat(timespec="seconds")
            else:
                if not custom_datetime:
                    flash("Please select a valid date and time.", "error")
                    return redirect(url_for("new_ticket"))
                try:
                    created_at = datetime.fromisoformat(custom_datetime).isoformat(timespec="seconds")
                except ValueError:
                    flash("Invalid date format.", "error")
                    return redirect(url_for("new_ticket"))

            row = {
                "ticket_number": ticket_number,
                "created_at": created_at,
                "direction": direction,
                "job_code_snapshot": job_code_snapshot,
                "job_name_snapshot": job_name_snapshot,
                "tax_exempt": tax_exempt_snapshot,
                "customer_snapshot": customer_snapshot,
                "truck_number_snapshot": truck_number_snapshot,
                "material_name_snapshot": material_name_snapshot,
                "quantity": quantity_num,
                "unit": unit,
                "cost": ticket_cost,
                "notes": notes,
            }
            pdf_bytes = to_pdf_bytes(row)
            pdf_path = save_pdf(ticket_number, pdf_bytes)
            app.logger.info("Generated PDF for ticket %s.", ticket_number)

            with db.cursor() as cursor:
                cursor.execute(
                """
                INSERT INTO tickets (
                    ticket_number, ticket_year, ticket_sequence, direction, created_at,
                    job_id, job_code_snapshot, job_name_snapshot, tax_exempt, customer_snapshot, truck_id, truck_number_snapshot,
                    material_id, material_name_snapshot, quantity, unit, cost, notes, pdf_path, pdf_blob
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    ticket_number,
                    ticket_year,
                    seq,
                    direction,
                    created_at,
                    job_id_value,
                    job_code_snapshot,
                    job_name_snapshot,
                    tax_exempt_snapshot,
                    customer_snapshot,
                    truck_id_value,
                    truck_number_snapshot,
                    material_id_value,
                    material_name_snapshot,
                    quantity_num,
                    unit,
                    ticket_cost,
                    notes,
                    pdf_path,
                    pdf_bytes,
                ),
                )
                inserted_ticket = cursor.fetchone()
                if inserted_ticket:
                    created_ticket_id = inserted_ticket["id"]
            db.commit()
        except Exception:
            db.rollback()
            raise

        image_result = None
        if created_ticket_id is not None:
            try:
                source_row = get_ticket_image_source_row(db, created_ticket_id)
                if source_row:
                    image_result = generate_ticket_image_for_row(db, source_row, force=False, capture_mode="live")
                    db.commit()
            except Exception as exc:
                db.rollback()
                app.logger.warning("Ticket %s image generation failed: %s", ticket_number, exc)
                image_result = {
                    "ok": False,
                    "code": "ERROR",
                    "message": str(exc),
                }

        if image_result and not image_result.get("ok"):
            msg = str(image_result.get("message") or "Could not generate ticket image.")
            if image_result.get("code") == "NO_FOOTAGE":
                flash("Ticket saved. No camera footage was available at ticket time.", "error")
            else:
                flash(f"Ticket saved. Image generation failed: {msg}", "error")

        if auto_print:
            if created_ticket_id is not None:
                # Use an intermediate print page: kiosk mode can print silently;
                # otherwise the user still gets a manual PDF fallback link.
                return redirect(url_for("ticket_auto_print", ticket_id=created_ticket_id))
            flash("Ticket saved, but PDF preview could not be opened automatically.", "error")
            return redirect(url_for("new_ticket"))

        flash(f"Ticket {ticket_number} created.", "success")
        return redirect(url_for("new_ticket"))
        # return redirect(url_for("search_tickets", ticket_number=ticket_number))

    else:
        direction = request.args.get("direction", "IN").strip().upper()

    return render_template(
        "ticket_new.html",
        jobs=list_ticket_jobs(db),
        customers =list_customers(db),
        trucks=list_trucks(db),
        materials=list_materials(db,direction=direction),
        recent_tickets=list_recent_tickets(db, limit=5),
    )

@app.get("/materials")
def get_materials():
    db = get_db()
    direction = request.args.get("direction")

    materials = list_materials(db, direction)

    return {
        "materials": [
            {"id": m["id"], "name": m["material_name"]}
            for m in materials
        ]
    }


@app.get("/tickets/prefill/latest")
def get_latest_ticket_prefill():
    db = get_db()
    truck_id_raw = request.args.get("truck_id", "").strip()
    truck_entry = request.args.get("truck_entry", "").strip()
    for_date_raw = request.args.get("for_date", "").strip()

    target_date = app_now().date()
    if for_date_raw:
        try:
            target_date = datetime.fromisoformat(for_date_raw).date()
        except ValueError:
            pass

    use_truck_id = truck_id_raw.isdigit()
    if not use_truck_id and not truck_entry:
        return {"found": False}

    query = """
        SELECT
            id,
            direction,
            job_id,
            job_code_snapshot,
            job_name_snapshot,
            customer_snapshot,
            material_id,
            material_name_snapshot,
            quantity,
            unit,
            notes
        FROM tickets
        WHERE COALESCE(active, TRUE) = TRUE
          AND date(created_at) = date(%s)
    """
    params = [target_date.isoformat()]

    if use_truck_id:
        query += " AND truck_id = %s"
        params.append(int(truck_id_raw))
    else:
        query += " AND truck_number_snapshot = %s"
        params.append(truck_entry)

    query += " ORDER BY created_at DESC, id DESC LIMIT 1"

    with db.cursor() as cursor:
        cursor.execute(query, tuple(params))
        row = cursor.fetchone()

    if not row:
        return {"found": False}

    customer_name = str(row.get("customer_snapshot") or "").strip()
    customer_id = None
    if customer_name:
        customer = get_customer_by_name(db, customer_name)
        if customer:
            customer_id = customer.get("id")

    job_code = str(row.get("job_code_snapshot") or "").strip()
    job_name = str(row.get("job_name_snapshot") or "").strip()
    job_entry = f"{job_code} - {job_name}" if job_code and job_name else (job_code or job_name)

    job_selected_key = ""
    if row.get("job_id"):
        job_selected_key = f"cache:{row['job_id']}"
    elif job_code or job_name:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id FROM manual_jobs WHERE job_code = %s AND job_name = %s LIMIT 1",
                (job_code, job_name),
            )
            manual_row = cursor.fetchone()
        if manual_row:
            job_selected_key = f"manual:{manual_row['id']}"
        else:
            with db.cursor() as cursor:
                cursor.execute(
                    "SELECT id FROM jobs_cache WHERE job_code = %s AND job_name = %s LIMIT 1",
                    (job_code, job_name),
                )
                cache_row = cursor.fetchone()
            if cache_row:
                job_selected_key = f"cache:{cache_row['id']}"

    return {
        "found": True,
        "direction": str(row.get("direction") or "").strip().upper(),
        "job_entry": job_entry,
        "job_selected_key": job_selected_key,
        "customer_name": customer_name,
        "customer_id": customer_id,
        "material_name": str(row.get("material_name_snapshot") or "").strip(),
        "material_id": row.get("material_id"),
        "quantity": row.get("quantity"),
        "unit": str(row.get("unit") or "").strip(),
        "notes": str(row.get("notes") or ""),
    }

@app.post("/jobs/refresh")
def refresh_jobs():
    db = get_db()
    try:
        count = refresh_jobs_cache(db)
        db.commit()
        flash(f"Jobs refresh complete. {count} rows synced.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Jobs refresh failed: {exc}", "error")
    return redirect(url_for("new_ticket"))


@app.get("/jobs/search")
def jobs_search():
    db = get_db()
    query_text = request.args.get("q", "")
    jobs = search_ticket_jobs(db, query_text)
    payload = [
        {
            "job_key": str(job.get("job_key") or "").strip(),
            "job_code": str(job.get("job_code") or "").strip(),
            "job_name": str(job.get("job_name") or "").strip(),
            "customer": str(job.get("customer") or "").strip(),
        }
        for job in jobs
    ]
    return {"jobs": payload}


@app.route("/tickets/search", methods=["GET"])
def search_tickets():
    db = get_db()
    ticket_number = request.args.get("ticket_number", "").strip()
    truck = request.args.get("truck", "").strip()
    job = request.args.get("job", "").strip()
    customer = request.args.get("customer", "").strip()
    material = request.args.get("material", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    query = """
        SELECT
            id,
            ticket_number,
            created_at,
            direction,
            job_code_snapshot,
            tax_exempt,
            customer_snapshot,
            truck_number_snapshot,
            material_name_snapshot,
            cost,
            image_url,
            image_status,
            image_error,
            CASE
                WHEN pdf_blob IS NOT NULL OR COALESCE(pdf_path, '') <> '' THEN TRUE
                ELSE FALSE
            END AS has_pdf
        FROM tickets
        WHERE COALESCE(active, TRUE) = TRUE
    """
    params = []
    if ticket_number:
        query += " AND ticket_number ILIKE %s"
        params.append(f"%{ticket_number}%")
    if truck:
        query += " AND truck_number_snapshot ILIKE %s"
        params.append(f"%{truck}%")
    if job:
        query += " AND job_code_snapshot ILIKE %s"
        params.append(f"%{job}%")
    if customer:
        query += " AND customer_snapshot ILIKE %s"
        params.append(f"%{customer}%")
    if material:
        query += " AND material_name_snapshot ILIKE %s"
        params.append(f"%{material}%")
    if date_from:
        query += " AND date(created_at) >= date(%s)"
        params.append(date_from)
    if date_to:
        query += " AND date(created_at) <= date(%s)"
        params.append(date_to)
    query += " ORDER BY created_at DESC, id DESC LIMIT 200"

    with db.cursor() as cursor:
        cursor.execute(query, tuple(params))
        tickets = cursor.fetchall()

    for ticket in tickets:
        ticket["image_view_url"] = build_ticket_image_view_url(ticket.get("image_url"))

    return render_template(
        "ticket_search.html",
        tickets=tickets,
        customers=list_customers(db),
    )


@app.route("/tickets/edit", methods=["GET"])
def edit_tickets():
    db = get_db()
    ticket_number = request.args.get("ticket_number", "").strip()
    truck = request.args.get("truck", "").strip()
    job = request.args.get("job", "").strip()
    customer = request.args.get("customer", "").strip()
    material = request.args.get("material", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    edit_id_raw = request.args.get("edit_id", "").strip()

    edit_id = None
    if edit_id_raw.isdigit():
        edit_id = int(edit_id_raw)

    # Default edit view to today's tickets when no explicit date range is provided.
    if not date_from and not date_to:
        today = app_now().date().isoformat()
        date_from = today
        date_to = today

    query = """
        SELECT
            id,
            ticket_number,
            created_at,
            direction,
            job_id,
            job_code_snapshot,
            job_name_snapshot,
            tax_exempt,
            customer_snapshot,
            truck_id,
            truck_number_snapshot,
            material_id,
            material_name_snapshot,
            quantity,
            unit,
            cost,
            notes
        FROM tickets
        WHERE COALESCE(active, TRUE) = TRUE
    """
    params = []
    if ticket_number:
        query += " AND ticket_number ILIKE %s"
        params.append(f"%{ticket_number}%")
    if truck:
        query += " AND truck_number_snapshot ILIKE %s"
        params.append(f"%{truck}%")
    if job:
        query += " AND job_code_snapshot ILIKE %s"
        params.append(f"%{job}%")
    if customer:
        query += " AND customer_snapshot ILIKE %s"
        params.append(f"%{customer}%")
    if material:
        query += " AND material_name_snapshot ILIKE %s"
        params.append(f"%{material}%")
    if date_from:
        query += " AND date(created_at) >= date(%s)"
        params.append(date_from)
    if date_to:
        query += " AND date(created_at) <= date(%s)"
        params.append(date_to)
    query += " ORDER BY created_at DESC, id DESC LIMIT 200"

    with db.cursor() as cursor:
        cursor.execute(query, tuple(params))
        tickets = cursor.fetchall()

    selected_ticket = None
    if edit_id is not None:
        selected_ticket = next((t for t in tickets if t.get("id") == edit_id), None)
        if not selected_ticket:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT
                        id,
                        ticket_number,
                        created_at,
                        direction,
                        job_id,
                        job_code_snapshot,
                        job_name_snapshot,
                        tax_exempt,
                        customer_snapshot,
                        truck_id,
                        truck_number_snapshot,
                        material_id,
                        material_name_snapshot,
                        quantity,
                        unit,
                        cost,
                        notes
                    FROM tickets
                    WHERE id = %s AND COALESCE(active, TRUE) = TRUE
                    """,
                    (edit_id,),
                )
                selected_ticket = cursor.fetchone()

    jobs = []
    customers = []
    trucks = []
    materials = []

    if selected_ticket:
        jobs = list_ticket_jobs(db)
        customers = list_customers(db)
        trucks = list_trucks(db)
        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    id,
                    material AS material_name,
                    direction,
                    axle_1,
                    tandem,
                    triaxle,
                    axle_4_5,
                    axle_6,
                    semi,
                    hydvac,
                    dirt_in
                FROM material_price
                WHERE active = TRUE
                ORDER BY material
                """
            )
            materials = cursor.fetchall()

        manual_lookup = {}
        with db.cursor() as cursor:
            cursor.execute("SELECT id, job_code, job_name FROM manual_jobs")
            for row in cursor.fetchall():
                manual_lookup[(str(row["job_code"] or "").strip(), str(row["job_name"] or "").strip())] = row["id"]

        cache_lookup = {}
        with db.cursor() as cursor:
            cursor.execute("SELECT id, job_code, job_name FROM jobs_cache")
            for row in cursor.fetchall():
                cache_lookup[(str(row["job_code"] or "").strip(), str(row["job_name"] or "").strip())] = row["id"]

        selected_ticket["created_at_input"] = to_datetime_local_value(selected_ticket.get("created_at"))
        job_code = str(selected_ticket.get("job_code_snapshot") or "").strip()
        job_name = str(selected_ticket.get("job_name_snapshot") or "").strip()
        if selected_ticket.get("job_id"):
            selected_ticket["job_selected"] = f"cache:{selected_ticket['job_id']}"
        elif (job_code, job_name) in manual_lookup:
            selected_ticket["job_selected"] = f"manual:{manual_lookup[(job_code, job_name)]}"
        elif (job_code, job_name) in cache_lookup:
            selected_ticket["job_selected"] = f"cache:{cache_lookup[(job_code, job_name)]}"
        else:
            selected_ticket["job_selected"] = ""

    return render_template(
        "tickets_edit.html",
        tickets=tickets,
        selected_ticket=selected_ticket,
        jobs=jobs,
        customers=customers,
        search_customers=list_customers(db),
        trucks=trucks,
        materials=materials,
        ticket_number=ticket_number,
        truck=truck,
        job=job,
        customer=customer,
        material=material,
        date_from=date_from,
        date_to=date_to,
        edit_id=edit_id,
    )


@app.post("/tickets/<int:ticket_id>/edit-save")
def save_ticket_edit(ticket_id):
    db = get_db()

    filter_keys = ["ticket_number", "truck", "job", "customer", "material", "date_from", "date_to", "edit_id"]
    filter_args = {}
    for key in filter_keys:
        value = request.form.get(key, "").strip()
        if value:
            filter_args[key] = value

    admin_password = request.form.get("admin_password", "")
    password_ok, password_message = validate_ticket_edit_password(admin_password)
    if not password_ok:
        flash(f"{password_message} Ticket was not updated.", "error")
        return redirect(url_for("edit_tickets", **filter_args))

    try:
        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, ticket_number
                FROM tickets
                WHERE id = %s AND COALESCE(active, TRUE) = TRUE
                """,
                (ticket_id,),
            )
            old_row = cursor.fetchone()

        if not old_row:
            flash("Ticket not found or already inactive.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        direction = request.form.get("direction", "").strip().upper()
        if direction not in {"IN", "OUT"}:
            flash("Direction must be IN or OUT.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        created_at_raw = request.form.get("created_at", "").strip()
        if not created_at_raw:
            flash("Date/Time is required.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        try:
            created_at = datetime.fromisoformat(created_at_raw).isoformat(timespec="seconds")
        except ValueError:
            flash("Invalid date/time format.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        selected_job_id = request.form.get("job_selected", "").strip()
        customer_id = parse_optional_int(request.form.get("customer_id", ""))
        truck_id = parse_optional_int(request.form.get("truck_id", ""))
        material_id = parse_optional_int(request.form.get("material_id", ""))

        if not selected_job_id or not customer_id or not truck_id or not material_id:
            flash("Please choose job, customer, truck, and material from dropdowns.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        selected_job, selected_job_source = get_selected_job(db, selected_job_id)
        if not selected_job:
            flash("Selected job was not found.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        with db.cursor() as cursor:
            cursor.execute("SELECT id, customer_name FROM customers WHERE id = %s", (customer_id,))
            selected_customer = cursor.fetchone()
        if not selected_customer:
            flash("Selected customer was not found.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        with db.cursor() as cursor:
            cursor.execute("SELECT id, truck_number, truck_size FROM trucks_main WHERE id = %s", (truck_id,))
            selected_truck = cursor.fetchone()
        if not selected_truck:
            flash("Selected truck was not found.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    id,
                    material AS material_name,
                    direction,
                    axle_1,
                    tandem,
                    triaxle,
                    axle_4_5,
                    axle_6,
                    semi,
                    hydvac,
                    dirt_in
                FROM material_price
                WHERE id = %s
                """,
                (material_id,),
            )
            selected_material = cursor.fetchone()
        if not selected_material:
            flash("Selected material was not found.", "error")
            return redirect(url_for("edit_tickets", **filter_args))
        if str(selected_material.get("direction") or "").strip().upper() != direction:
            flash("Selected material direction does not match ticket direction.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        job_id_value = selected_job["id"] if selected_job_source == "cache" else None
        job_code_snapshot = str(selected_job.get("job_code") or "").strip()
        job_name_snapshot = str(selected_job.get("job_name") or "").strip()
        tax_exempt = str(selected_job.get("tax_exempt") or "").strip()
        customer_snapshot = str(selected_customer.get("customer_name") or "").strip()
        truck_number_snapshot = str(selected_truck.get("truck_number") or "").strip()
        material_name_snapshot = str(selected_material.get("material_name") or "").strip()
        unit = request.form.get("unit", "").strip()
        notes = request.form.get("notes", "").strip()

        if not all([
            job_code_snapshot,
            job_name_snapshot,
            customer_snapshot,
            truck_number_snapshot,
            material_name_snapshot,
            unit,
        ]):
            flash("Job, customer, truck, material, and unit fields are required.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        try:
            quantity = float(request.form.get("quantity", "").strip())
        except ValueError:
            flash("Quantity must be numeric.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        cost = calculate_ticket_cost(selected_truck, selected_material, quantity)

        ticket_number = str(old_row.get("ticket_number") or "").strip()
        if not ticket_number:
            flash("Ticket number is missing on this record.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        updated_ticket_row = {
            "ticket_number": ticket_number,
            "created_at": created_at,
            "direction": direction,
            "job_code_snapshot": job_code_snapshot,
            "job_name_snapshot": job_name_snapshot,
            "tax_exempt": tax_exempt,
            "customer_snapshot": customer_snapshot,
            "truck_number_snapshot": truck_number_snapshot,
            "material_name_snapshot": material_name_snapshot,
            "quantity": quantity,
            "unit": unit,
            "cost": cost,
            "notes": notes,
        }

        pdf_bytes = to_pdf_bytes(updated_ticket_row)
        pdf_path = save_pdf(ticket_number, pdf_bytes)

        with db.cursor() as cursor:
            cursor.execute(
                """
                UPDATE tickets
                SET
                    direction = %s,
                    created_at = %s,
                    job_id = %s,
                    job_code_snapshot = %s,
                    job_name_snapshot = %s,
                    tax_exempt = %s,
                    customer_snapshot = %s,
                    truck_id = %s,
                    truck_number_snapshot = %s,
                    material_id = %s,
                    material_name_snapshot = %s,
                    quantity = %s,
                    unit = %s,
                    cost = %s,
                    notes = %s,
                    active = TRUE,
                    modified_to_id = NULL,
                    pdf_path = %s,
                    pdf_blob = %s
                WHERE id = %s
                """,
                (
                    direction,
                    created_at,
                    job_id_value,
                    job_code_snapshot,
                    job_name_snapshot,
                    tax_exempt,
                    customer_snapshot,
                    selected_truck["id"],
                    truck_number_snapshot,
                    selected_material["id"],
                    material_name_snapshot,
                    quantity,
                    unit,
                    cost,
                    notes,
                    pdf_path,
                    pdf_bytes,
                    ticket_id,
                ),
            )

        db.commit()
        flash(f"Ticket {ticket_number} updated.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Could not update ticket: {exc}", "error")

    return redirect(url_for("edit_tickets", **filter_args))


@app.post("/tickets/<int:ticket_id>/edit-void")
def void_ticket_from_edit(ticket_id):
    db = get_db()

    filter_keys = ["ticket_number", "truck", "job", "customer", "material", "date_from", "date_to", "edit_id"]
    filter_args = {}
    for key in filter_keys:
        value = request.form.get(key, "").strip()
        if value:
            filter_args[key] = value

    admin_password = request.form.get("admin_password", "")
    password_ok, password_message = validate_ticket_edit_password(admin_password)
    if not password_ok:
        flash(f"{password_message} Ticket was not voided.", "error")
        return redirect(url_for("edit_tickets", **filter_args))

    try:
        with db.cursor() as cursor:
            cursor.execute(
                """
                UPDATE tickets
                SET active = FALSE
                WHERE id = %s AND COALESCE(active, TRUE) = TRUE
                RETURNING ticket_number
                """,
                (ticket_id,),
            )
            row = cursor.fetchone()

        if not row:
            db.rollback()
            flash("Ticket not found or already inactive.", "error")
            return redirect(url_for("edit_tickets", **filter_args))

        db.commit()
        flash(f"Ticket {row['ticket_number']} was voided.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Could not void ticket: {exc}", "error")

    return redirect(url_for("edit_tickets", **filter_args))


@app.route("/reports", methods=["GET"])
def reports():
    db = get_db()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    direction = request.args.get("direction", "").strip().upper()
    job_id = request.args.get("job_id", "").strip()
    job_entry = request.args.get("job_entry", "").strip()
    customer_id = request.args.get("customer_id", "").strip()
    customer = request.args.get("customer", "").strip()
    material_id = request.args.get("material_id", "").strip()
    material_entry = request.args.get("material_entry", "").strip()
    offset = int(request.args.get("offset", 0))
    if not date_from and not date_to:
        date_to = app_now().date().isoformat()
        date_from = (app_now().date() - timedelta(days=14)).isoformat()

    where = ["COALESCE(t.active, TRUE) = TRUE"]
    params = []

    if date_from:
        where.append("date(t.created_at) >= date(%s)")
        params.append(date_from)
    if date_to:
        where.append("date(t.created_at) <= date(%s)")
        params.append(date_to)
    if direction in {"IN", "OUT"}:
        where.append("t.direction = %s")
        params.append(direction)
    apply_reports_job_filter(db, where, params, selected_job_id=job_id, job_entry=job_entry)
    apply_reports_customer_filter(db, where, params, customer_id=customer_id, customer=customer)
    apply_reports_material_filter(where, params, material_id=material_id, material_entry=material_entry)

    where_sql = " AND ".join(where)
    app.logger.debug("Report query WHERE clause: %s with params %s", where_sql, params)

    with db.cursor() as cursor:
        cursor.execute(
        f"""
        SELECT
            t.id,
            t.ticket_number,
            t.created_at,
            t.direction,
            t.job_code_snapshot,
            t.job_name_snapshot,
            t.customer_snapshot,
            t.tax_exempt,
            t.material_name_snapshot,
            t.truck_number_snapshot,
            t.quantity,
            t.unit,
            t.cost
        FROM tickets t
        WHERE {where_sql}
        ORDER BY 
            t.customer_snapshot ASC,
            CASE 
                WHEN t.direction = 'IN' THEN 1
                WHEN t.direction = 'OUT' THEN 2
                ELSE 3 
            END,
            t.created_at DESC,
            t.id DESC
        LIMIT 20 OFFSET %s
        """,
        tuple(params+[offset]),
        )
        tickets = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT t.unit,
             COALESCE(SUM(t.quantity), 0) AS total_quantity,
             COALESCE(SUM(t.cost), 0) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY t.unit
        ORDER BY t.unit
        """,
        tuple(params),
        )
        totals_by_unit = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT t.material_name_snapshot,
             t.unit,
             COALESCE(SUM(t.quantity), 0) AS total_quantity,
             COALESCE(SUM(t.cost), 0) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY t.material_name_snapshot, t.unit
        ORDER BY t.material_name_snapshot, t.unit
        """,
        tuple(params),
        )
        totals_by_material = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT t.direction,
             t.unit,
             COALESCE(SUM(t.quantity), 0) AS total_quantity,
             COALESCE(SUM(t.cost), 0) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY t.direction, t.unit
        ORDER BY t.direction, t.unit
        """,
        tuple(params),
        )
        totals_by_direction = cursor.fetchall()

    if not job_entry and job_id:
        job_entry = get_job_entry_display(db, job_id)
    if not customer and customer_id:
        customer = get_customer_display_by_id(db, customer_id)
    if not material_entry and material_id:
        material_entry = get_material_entry_display(db, material_id)

    return render_template(
        "reports.html",
        tickets=tickets,
        offset=offset,
        today_date=app_now().date().isoformat(),
        totals_by_unit=totals_by_unit,
        totals_by_material=totals_by_material,
        totals_by_direction=totals_by_direction,
        jobs=list_ticket_jobs(db),
        customers=list_customers(db),
        materials=list_materials(db,direction=direction),
        filters={
            "date_from": date_from,
            "date_to": date_to,
            "direction": direction,
            "job_id": job_id,
            "job_entry": job_entry,
            "customer_id": customer_id,
            "customer": customer,
            "material_id": material_id,
            "material_entry": material_entry,
        },
    )


@app.route("/reports/export.csv", methods=["GET"])
def export_reports_csv():
    db = get_db()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    direction = request.args.get("direction", "").strip().upper()
    job_id = request.args.get("job_id", "").strip()
    job_entry = request.args.get("job_entry", "").strip()
    customer_id = request.args.get("customer_id", "").strip()
    customer = request.args.get("customer", "").strip()
    material_id = request.args.get("material_id", "").strip()
    material_entry = request.args.get("material_entry", "").strip()

    where = ["COALESCE(t.active, TRUE) = TRUE"]
    params = []

    if date_from:
        where.append("date(t.created_at) >= date(%s)")
        params.append(date_from)
    if date_to:
        where.append("date(t.created_at) <= date(%s)")
        params.append(date_to)
    if direction in {"IN", "OUT"}:
        where.append("t.direction = %s")
        params.append(direction)
    apply_reports_job_filter(db, where, params, selected_job_id=job_id, job_entry=job_entry)
    apply_reports_customer_filter(db, where, params, customer_id=customer_id, customer=customer)
    apply_reports_material_filter(where, params, material_id=material_id, material_entry=material_entry)

    where_sql = " AND ".join(where)

    with db.cursor() as cursor:
        cursor.execute(
        f"""
        SELECT
            t.ticket_number,
            t.created_at,
            t.direction,
            t.job_code_snapshot,
            t.job_name_snapshot,
            t.customer_snapshot,
            t.truck_number_snapshot,
            t.material_name_snapshot,
            t.quantity,
            t.unit,
            t.cost
        FROM tickets t
        WHERE {where_sql}
        ORDER BY t.created_at DESC, t.id DESC
        LIMIT 1000
        """,
        tuple(params),
        )
        tickets = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT t.unit,
             COALESCE(SUM(t.quantity), 0) AS total_quantity,
             COALESCE(SUM(t.cost), 0) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY t.unit
        ORDER BY t.unit
        """,
        tuple(params),
        )
        totals_by_unit = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT t.material_name_snapshot,
             t.unit,
             COALESCE(SUM(t.quantity), 0) AS total_quantity,
             COALESCE(SUM(t.cost), 0) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY t.material_name_snapshot, t.unit
        ORDER BY t.material_name_snapshot, t.unit
        """,
        tuple(params),
        )
        totals_by_material = cursor.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        [
            "Ticket Number",
            "Created At",
            "Direction",
            "Job Code",
            "Job Name",
            "Customer",
            "Truck",
            "Material",
            "Quantity",
            "Unit",
            "Cost",
        ]
    )
    for t in tickets:
        writer.writerow(
            [
                t["ticket_number"],
                format_ticket_datetime(t["created_at"]),
                t["direction"],
                t["job_code_snapshot"],
                t["job_name_snapshot"],
                t["customer_snapshot"],
                t["truck_number_snapshot"],
                t["material_name_snapshot"],
                f"{t['quantity']:.2f}",
                t["unit"],
                f"{float(t.get('cost', 0) or 0):.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["Totals by Unit"])
    writer.writerow(["Unit", "Total Quantity", "Total Cost"])
    for total in totals_by_unit:
        writer.writerow([total["unit"], f"{total['total_quantity']:.2f}", f"{float(total.get('total_cost', 0) or 0):.2f}"])

    writer.writerow([])
    writer.writerow(["Totals by Material"])
    writer.writerow(["Material", "Unit", "Total Quantity", "Total Cost"])
    for total in totals_by_material:
        writer.writerow(
            [
                total["material_name_snapshot"],
                total["unit"],
                f"{total['total_quantity']:.2f}",
                f"{float(total.get('total_cost', 0) or 0):.2f}",
            ]
        )

    csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))
    output.close()
    csv_bytes.seek(0)

    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    download_filename = f"ticket_report_{stamp}.csv"
    csv_payload = csv_bytes.getvalue()
    try:
        upload_download_audit_blob(
            category="reports_csv",
            filename=download_filename,
            file_bytes=csv_payload,
            mimetype="text/csv",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload report CSV download: %s", exc)

    return send_file(
        io.BytesIO(csv_payload),
        mimetype="text/csv",
        as_attachment=True,
        download_name=download_filename,
    )


@app.post("/reports/credit-card/daily")
@app.post("/api/reports/credit-card/daily")
def api_credit_card_daily_report():
    expected_key = CREDIT_CARD_REPORT_API_KEY
    provided_key = (request.headers.get("X-API-Key") or request.args.get("api_key") or "").strip()

    if not expected_key:
        return {"ok": False, "error": "CREDIT_CARD_REPORT_API_KEY is not configured."}, 503
    if not hmac.compare_digest(provided_key, expected_key):
        return {"ok": False, "error": "Unauthorized."}, 401

    body = request.get_json(silent=True) if request.is_json else {}
    body = body or {}
    if not body:
        raw_body = request.get_data(as_text=True) or ""
        if raw_body.strip():
            try:
                parsed_body = json.loads(raw_body)
                if isinstance(parsed_body, dict):
                    body = parsed_body
            except Exception:
                pass

    report_date_raw = str(
        body.get("report_date")
        or request.form.get("report_date")
        or request.args.get("report_date")
        or ""
    ).strip()
    sas_minutes_raw = str(
        body.get("sas_minutes")
        or request.form.get("sas_minutes")
        or request.args.get("sas_minutes")
        or CREDIT_CARD_REPORT_SAS_MINUTES
    ).strip()

    report_date = app_now().date()
    if report_date_raw:
        try:
            report_date = datetime.fromisoformat(report_date_raw).date()
        except ValueError:
            return {"ok": False, "error": "Invalid report_date. Use YYYY-MM-DD."}, 400

    try:
        sas_minutes = int(sas_minutes_raw)
    except ValueError:
        return {"ok": False, "error": "Invalid sas_minutes. Use an integer."}, 400

    db = get_db()
    rows = fetch_credit_card_sales_rows(db, report_date, CREDIT_CARD_CUSTOMER_MATCH)
    pdf_bytes = credit_card_daily_report_to_pdf_bytes(rows, report_date, CREDIT_CARD_CUSTOMER_MATCH)
    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    filename = f"credit_card_sales_{report_date.strftime('%Y%m%d')}_{stamp}.pdf"

    if not AZURE_STORAGE_CONNECTION_STRING:
        return {"ok": False, "error": "AZURE_STORAGE_CONNECTION_STRING is not configured."}, 503

    blob_name = f"{AZURE_REPORTS_BLOB_PREFIX}/{filename}" if AZURE_REPORTS_BLOB_PREFIX else filename
    try:
        blob_url = upload_pdf_to_blob(blob_name, pdf_bytes)
    except Exception as exc:
        return {"ok": False, "error": f"Blob upload failed: {exc}"}, 500

    if not blob_url:
        return {"ok": False, "error": "Blob upload failed: no blob URL returned."}, 500

    try:
        sas_url, sas_expires_at_utc = generate_blob_read_sas_url(blob_name, sas_minutes)
    except Exception as exc:
        return {"ok": False, "error": f"SAS generation failed: {exc}"}, 500

    try:
        upload_download_audit_blob(
            category="reports_pdf_credit_card",
            filename=filename,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload API credit-card report PDF: %s", exc)

    return {
        "ok": True,
        "report_date": report_date.isoformat(),
        "customer_match": CREDIT_CARD_CUSTOMER_MATCH,
        "transactions": len(rows),
        "total_cost": round(sum(float(r.get("cost") or 0) for r in rows), 2),
        "pdf_filename": filename,
        "blob_url": blob_url,
        "sas_url": sas_url,
        "sas_expires_at_utc": sas_expires_at_utc,
        "sas_minutes": sas_minutes,
    }, 200


@app.post("/reports/non-credit-card/daily")
@app.post("/api/reports/non-credit-card/daily")
def api_non_credit_card_daily_report():
    expected_key = CREDIT_CARD_REPORT_API_KEY
    provided_key = (request.headers.get("X-API-Key") or request.args.get("api_key") or "").strip()

    if not expected_key:
        return {"ok": False, "error": "CREDIT_CARD_REPORT_API_KEY is not configured."}, 503
    if not hmac.compare_digest(provided_key, expected_key):
        return {"ok": False, "error": "Unauthorized."}, 401

    body = request.get_json(silent=True) if request.is_json else {}
    body = body or {}
    if not body:
        raw_body = request.get_data(as_text=True) or ""
        if raw_body.strip():
            try:
                parsed_body = json.loads(raw_body)
                if isinstance(parsed_body, dict):
                    body = parsed_body
            except Exception:
                pass

    report_date_raw = str(
        body.get("report_date")
        or request.form.get("report_date")
        or request.args.get("report_date")
        or ""
    ).strip()
    sas_minutes_raw = str(
        body.get("sas_minutes")
        or request.form.get("sas_minutes")
        or request.args.get("sas_minutes")
        or CREDIT_CARD_REPORT_SAS_MINUTES
    ).strip()

    report_date = app_now().date()
    if report_date_raw:
        try:
            report_date = datetime.fromisoformat(report_date_raw).date()
        except ValueError:
            return {"ok": False, "error": "Invalid report_date. Use YYYY-MM-DD."}, 400

    try:
        sas_minutes = int(sas_minutes_raw)
    except ValueError:
        return {"ok": False, "error": "Invalid sas_minutes. Use an integer."}, 400

    db = get_db()
    week_start = report_date - timedelta(days=report_date.weekday())
    week_end = week_start + timedelta(days=6)

    rows = fetch_non_credit_card_sales_rows(db, week_start, week_end, CREDIT_CARD_CUSTOMER_MATCH)
    pdf_bytes = non_credit_card_daily_report_to_pdf_bytes(rows, week_start, week_end, CREDIT_CARD_CUSTOMER_MATCH)
    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    filename = f"non_credit_card_sales_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}_{stamp}.pdf"

    if not AZURE_STORAGE_CONNECTION_STRING:
        return {"ok": False, "error": "AZURE_STORAGE_CONNECTION_STRING is not configured."}, 503

    blob_name = f"{AZURE_REPORTS_BLOB_PREFIX}/{filename}" if AZURE_REPORTS_BLOB_PREFIX else filename
    try:
        blob_url = upload_pdf_to_blob(blob_name, pdf_bytes)
    except Exception as exc:
        return {"ok": False, "error": f"Blob upload failed: {exc}"}, 500

    if not blob_url:
        return {"ok": False, "error": "Blob upload failed: no blob URL returned."}, 500

    try:
        sas_url, sas_expires_at_utc = generate_blob_read_sas_url(blob_name, sas_minutes)
    except Exception as exc:
        return {"ok": False, "error": f"SAS generation failed: {exc}"}, 500

    try:
        upload_download_audit_blob(
            category="reports_pdf_non_credit_card",
            filename=filename,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload API non-credit-card report PDF: %s", exc)

    customers = sorted({str(r.get("customer_snapshot") or "").strip() or "(No Customer)" for r in rows}, key=lambda x: x.lower())
    return {
        "ok": True,
        "report_date": report_date.isoformat(),
        "week_start_date": week_start.isoformat(),
        "week_end_date": week_end.isoformat(),
        "excluded_customer_match": CREDIT_CARD_CUSTOMER_MATCH,
        "customers": customers,
        "customer_count": len(customers),
        "transactions": len(rows),
        "total_cost": round(sum(float(r.get("cost") or 0) for r in rows), 2),
        "pdf_filename": filename,
        "blob_url": blob_url,
        "sas_url": sas_url,
        "sas_expires_at_utc": sas_expires_at_utc,
        "sas_minutes": sas_minutes,
    }, 200


@app.get("/reports/print")
def print_reports():
    db = get_db()

    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    direction = request.args.get("direction", "")
    job_id = request.args.get("job_id", "")
    job_entry = request.args.get("job_entry", "").strip()
    customer_id = request.args.get("customer_id", "").strip()
    customer = request.args.get("customer", "").strip()
    material_id = request.args.get("material_id", "")
    material_entry = request.args.get("material_entry", "").strip()

    where = ["COALESCE(t.active, TRUE)=TRUE"]
    params = []

    if date_from:
        where.append("date(t.created_at)>=date(%s)")
        params.append(date_from)

    if date_to:
        where.append("date(t.created_at)<=date(%s)")
        params.append(date_to)

    if direction in {"IN", "OUT"}:
        where.append("t.direction=%s")
        params.append(direction)

    apply_reports_job_filter(db, where, params, selected_job_id=job_id, job_entry=job_entry)
    apply_reports_customer_filter(db, where, params, customer_id=customer_id, customer=customer)
    apply_reports_material_filter(where, params, material_id=material_id, material_entry=material_entry)

    if not customer and customer_id:
        customer = get_customer_display_by_id(db, customer_id)
    if not material_entry and material_id:
        material_entry = get_material_entry_display(db, material_id)

    where_sql = " AND ".join(where)

    with db.cursor() as cursor:
        cursor.execute(
        f"""
        SELECT *
        FROM tickets t
        WHERE {where_sql}
        ORDER BY
            t.customer_snapshot ASC,
            CASE
                WHEN t.direction='IN' THEN 1
                WHEN t.direction='OUT' THEN 2
            END,
            t.created_at DESC,
            t.id DESC
        """,
        tuple(params),
        )
        tickets = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT unit,
             SUM(quantity) AS total_quantity,
             SUM(cost) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY unit
        """,
        tuple(params),
        )
        totals_by_unit = cursor.fetchall()

    with db.cursor() as cursor:
        cursor.execute(
        f"""
         SELECT material_name_snapshot,
             unit,
             SUM(quantity) AS total_quantity,
             SUM(cost) AS total_cost
        FROM tickets t
        WHERE {where_sql}
        GROUP BY material_name_snapshot, unit
        """,
        tuple(params),
        )
        totals_by_material = cursor.fetchall()

    pdf_bytes = report_to_pdf_bytes(
        tickets,
        totals_by_unit,
        totals_by_material,
        {
            "date_from": date_from,
            "date_to": date_to,
            "direction": direction,
            "job_id": job_id,
            "customer": customer,
            "material_id": material_id,
            "material_entry": material_entry,
        },
    )

    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"ticket_report_{stamp}.pdf"

    if AZURE_STORAGE_CONNECTION_STRING:
        blob_name = f"{AZURE_REPORTS_BLOB_PREFIX}/{report_filename}" if AZURE_REPORTS_BLOB_PREFIX else report_filename
        try:
            upload_pdf_to_blob(blob_name, pdf_bytes)
        except Exception as exc:
            app.logger.warning("Azure Blob upload failed for report %s: %s", report_filename, exc)

        if os.name == "nt":
            try:
                temp_print_path = write_temp_pdf_for_print(pdf_bytes, "report")
                print_pdf_file(temp_print_path)
            except Exception:
                pass

        try:
            upload_download_audit_blob(
                category="reports_pdf",
                filename=report_filename,
                file_bytes=pdf_bytes,
                mimetype="application/pdf",
            )
        except Exception as exc:
            app.logger.warning("Could not audit-upload report PDF download: %s", exc)

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=report_filename,
        )

    pdf_path = REPORT_PDF_DIR / report_filename
    with open(pdf_path, "wb") as f:
        f.write(pdf_bytes)

    try:
        print_pdf_file(str(pdf_path))
    except Exception:
        pass

    try:
        upload_download_audit_blob(
            category="reports_pdf",
            filename=pdf_path.name,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload local report PDF download: %s", exc)

    return send_file(
        pdf_path,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=pdf_path.name,
    )


@app.get("/reports/print-customer")
def print_customer_reports():
    db = get_db()

    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    direction = request.args.get("direction", "")
    job_id = request.args.get("job_id", "")
    job_entry = request.args.get("job_entry", "").strip()
    customer_id = request.args.get("customer_id", "").strip()
    customer = request.args.get("customer", "").strip()
    material_id = request.args.get("material_id", "")
    material_entry = request.args.get("material_entry", "").strip()

    where = ["COALESCE(t.active, TRUE)=TRUE"]
    params = []

    if date_from:
        where.append("date(t.created_at)>=date(%s)")
        params.append(date_from)

    if date_to:
        where.append("date(t.created_at)<=date(%s)")
        params.append(date_to)

    if direction in {"IN", "OUT"}:
        where.append("t.direction=%s")
        params.append(direction)

    apply_reports_job_filter(db, where, params, selected_job_id=job_id, job_entry=job_entry)
    apply_reports_customer_filter(db, where, params, customer_id=customer_id, customer=customer)
    apply_reports_material_filter(where, params, material_id=material_id, material_entry=material_entry)

    where_sql = " AND ".join(where)

    with db.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                t.ticket_number,
                t.created_at,
                t.customer_snapshot,
                t.job_code_snapshot,
                t.job_name_snapshot,
                t.truck_number_snapshot,
                t.material_name_snapshot,
                t.cost
            FROM tickets t
            WHERE {where_sql}
            ORDER BY COALESCE(NULLIF(TRIM(t.customer_snapshot), ''), 'zzzzzz') ASC, t.created_at ASC, t.id ASC
            """,
            tuple(params),
        )
        rows = cursor.fetchall()

    job_label = ""
    if job_id:
        job_label = get_job_entry_display(db, job_id)
    if not job_label and job_entry:
        job_label = job_entry

    material_label = ""
    if material_id:
        material_label = get_material_entry_display(db, material_id)
    if not material_label and material_entry:
        material_label = material_entry

    pdf_bytes = customer_grouped_report_to_pdf_bytes(
        rows,
        {
            "date_from": date_from,
            "date_to": date_to,
            "direction": direction,
            "job_label": job_label,
            "customer": customer,
            "material_label": material_label,
        },
    )

    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"customer_grouped_report_{stamp}.pdf"

    if AZURE_STORAGE_CONNECTION_STRING:
        blob_name = f"{AZURE_REPORTS_BLOB_PREFIX}/{report_filename}" if AZURE_REPORTS_BLOB_PREFIX else report_filename
        try:
            upload_pdf_to_blob(blob_name, pdf_bytes)
        except Exception as exc:
            app.logger.warning("Azure Blob upload failed for customer report %s: %s", report_filename, exc)

    try:
        upload_download_audit_blob(
            category="reports_pdf_customer_grouped",
            filename=report_filename,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload customer grouped report PDF download: %s", exc)

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=report_filename,
    )


@app.get("/reports/daily-print")
def print_daily_report():
    db = get_db()
    report_date_raw = request.args.get("report_date", "").strip()
    report_date = app_now().date()
    if report_date_raw:
        try:
            report_date = datetime.fromisoformat(report_date_raw).date()
        except ValueError:
            pass
    report_date_str = report_date.isoformat()

    job_blocks, totals = build_daily_report_data(db, report_date_str)

    pdf_bytes = daily_report_to_pdf_bytes(
        job_blocks=job_blocks,
        report_date=report_date,
        totals=totals,
    )

    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"daily_report_{report_date.strftime('%Y%m%d')}_{stamp}.pdf"

    if AZURE_STORAGE_CONNECTION_STRING:
        blob_name = f"{AZURE_REPORTS_BLOB_PREFIX}/{report_filename}" if AZURE_REPORTS_BLOB_PREFIX else report_filename
        try:
            upload_pdf_to_blob(blob_name, pdf_bytes)
        except Exception as exc:
            app.logger.warning("Azure Blob upload failed for daily report %s: %s", report_filename, exc)

    try:
        upload_download_audit_blob(
            category="reports_pdf_daily",
            filename=report_filename,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload daily report PDF download: %s", exc)

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=report_filename,
    )


@app.get("/reports/daily")
def daily_report_dashboard():
    db = get_db()
    report_date_raw = request.args.get("report_date", "").strip()
    report_date = app_now().date()
    if report_date_raw:
        try:
            report_date = datetime.fromisoformat(report_date_raw).date()
        except ValueError:
            pass
    report_date_str = report_date.isoformat()

    job_blocks, totals = build_daily_report_data(db, report_date_str)

    return render_template(
        "daily_report.html",
        report_date=report_date.strftime("%m/%d/%Y"),
        report_date_input=report_date_str,
        job_blocks=job_blocks,
        totals=totals,
    )

@app.get("/tickets/<int:ticket_id>/pdf")
def ticket_pdf(ticket_id):
    db = get_db()
    inline_pdf = str(request.args.get("inline", "")).strip().lower() in {"1", "true", "yes", "on"}
    with db.cursor() as cursor:
        cursor.execute("SELECT ticket_number, pdf_path, pdf_blob FROM tickets WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()
    if not row:
        flash("Ticket not found.", "error")
        return redirect(url_for("search_tickets"))

    ticket_filename = f"{row['ticket_number']}.pdf"
    pdf_blob = row.get("pdf_blob")
    if pdf_blob is not None:
        try:
            upload_download_audit_blob(
                category="ticket_pdf",
                filename=ticket_filename,
                file_bytes=pdf_blob,
                mimetype="application/pdf",
            )
        except Exception as exc:
            app.logger.warning("Could not audit-upload ticket PDF download: %s", exc)

        return send_file(
            io.BytesIO(pdf_blob),
            mimetype="application/pdf",
            as_attachment=not inline_pdf,
            download_name=ticket_filename,
        )

    pdf_path = str(row.get("pdf_path") or "").strip()
    if pdf_path and not pdf_path.lower().startswith("http"):
        path_obj = Path(pdf_path)
        if path_obj.exists():
            return send_file(
                path_obj,
                mimetype="application/pdf",
                as_attachment=not inline_pdf,
                download_name=ticket_filename,
            )

    flash("No PDF exists for this ticket yet. Use Generate PDF.", "error")
    return redirect(url_for("search_tickets", ticket_number=row["ticket_number"]))


@app.get("/tickets/<int:ticket_id>/auto-print")
def ticket_auto_print(ticket_id):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT ticket_number FROM tickets WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()

    if not row:
        flash("Ticket not found.", "error")
        return redirect(url_for("new_ticket"))

    return render_template(
        "ticket_auto_print.html",
        ticket_id=ticket_id,
        ticket_number=row["ticket_number"],
        pdf_url=url_for("ticket_pdf", ticket_id=ticket_id, inline=1),
    )


@app.post("/tickets/<int:ticket_id>/print")
def print_ticket(ticket_id):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT ticket_number, pdf_path, pdf_blob FROM tickets WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()
    if not row:
        flash("Ticket not found.", "error")
        return redirect(url_for("search_tickets"))

    try:
        pdf_path = (row.get("pdf_path") or "").strip()
        pdf_blob = row.get("pdf_blob")
        if pdf_path and not pdf_path.lower().startswith("http") and Path(pdf_path).exists():
            print_pdf_file(pdf_path)
        elif pdf_blob is not None:
            temp_print_path = write_temp_pdf_for_print(pdf_blob, row["ticket_number"])
            print_pdf_file(temp_print_path)
        else:
            flash("No PDF exists for this ticket yet. Use Generate PDF first.", "error")
            return redirect(url_for("search_tickets", ticket_number=row["ticket_number"]))
        flash(f"Print sent for {row['ticket_number']}.", "success")
    except Exception as exc:
        flash(f"Print failed: {exc}", "error")
    return redirect(url_for("search_tickets", ticket_number=row["ticket_number"]))


@app.post("/tickets/<int:ticket_id>/generate-pdf")
def generate_ticket_pdf(ticket_id):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                id,
                ticket_number,
                created_at,
                direction,
                job_code_snapshot,
                job_name_snapshot,
                tax_exempt,
                customer_snapshot,
                truck_number_snapshot,
                material_name_snapshot,
                quantity,
                unit,
                cost,
                notes,
                pdf_path,
                pdf_blob
            FROM tickets
            WHERE id = %s
            """,
            (ticket_id,),
        )
        row = cursor.fetchone()

    if not row:
        flash("Ticket not found.", "error")
        return redirect(url_for("search_tickets"))

    try:
        pdf_bytes = to_pdf_bytes(row)
        pdf_path = save_pdf(row["ticket_number"], pdf_bytes)
        with db.cursor() as cursor:
            cursor.execute(
                "UPDATE tickets SET pdf_path = %s, pdf_blob = %s WHERE id = %s",
                (pdf_path, pdf_bytes, ticket_id),
            )
        db.commit()
        flash(f"PDF generated for {row['ticket_number']}.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Could not generate PDF: {exc}", "error")

    return redirect(url_for("search_tickets", ticket_number=row["ticket_number"]))


@app.post("/tickets/<int:ticket_id>/generate-image")
@app.post("/api/tickets/<int:ticket_id>/generate-image")
def generate_ticket_image(ticket_id):
    db = get_db()
    row = get_ticket_image_source_row(db, ticket_id)
    if not row:
        if request.path.startswith("/api/"):
            return {"ok": False, "error": "Ticket not found."}, 404
        flash("Ticket not found.", "error")
        return redirect(request.referrer or url_for("search_tickets"))

    try:
        result = generate_ticket_image_for_row(db, row, force=False, capture_mode="historical")
        db.commit()
    except Exception as exc:
        db.rollback()
        result = {"ok": False, "code": "ERROR", "message": str(exc)}

    if request.path.startswith("/api/"):
        status_code = 200 if result.get("ok") else 500
        if result.get("code") == "NO_FOOTAGE":
            status_code = 404
        if result.get("code") in {"WAVE_NOT_CONFIGURED", "DISABLED"}:
            status_code = 503
        payload = {
            "ok": bool(result.get("ok")),
            "ticket_id": ticket_id,
            "ticket_number": row.get("ticket_number"),
            "image_url": build_ticket_image_view_url(
                result.get("image_url") or str(row.get("image_url") or "").strip() or None
            ) or None,
            "code": result.get("code") or ("READY" if result.get("ok") else "ERROR"),
            "message": result.get("message") or "",
            "already_exists": bool(result.get("already_exists")),
        }
        return payload, status_code

    if result.get("ok"):
        if result.get("already_exists"):
            flash(f"Image already exists for {row['ticket_number']}.", "success")
        else:
            flash(f"Image generated for {row['ticket_number']}.", "success")
    else:
        message = str(result.get("message") or "Could not generate image.")
        if result.get("code") == "NO_FOOTAGE":
            flash(f"No footage for {row['ticket_number']} at that timestamp.", "error")
        else:
            flash(f"Image generation failed for {row['ticket_number']}: {message}", "error")

    return redirect(request.referrer or url_for("search_tickets", ticket_number=row["ticket_number"]))


@app.get("/rfid")
def rfid_home():
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT COUNT(1) AS count FROM rfid_tag_catalog")
        tags_count = int((cursor.fetchone() or {}).get("count") or 0)
        cursor.execute("SELECT COUNT(1) AS count FROM rfid_truck_tag_assignment")
        assigned_count = int((cursor.fetchone() or {}).get("count") or 0)
    return render_template("rfid_home.html", tags_count=tags_count, assigned_count=assigned_count)


@app.get("/rfid/mappings")
def rfid_mappings():
    db = get_db()
    assignments = list_rfid_assignments(db)
    unassigned_tags = list_unassigned_rfid_tags(db)
    unassigned_trucks = list_unassigned_trucks(db)

    trucks_for_edit = {}
    for assignment in assignments:
        current_truck = {
            "id": assignment.get("truck_id"),
            "truck_number": assignment.get("truck_number"),
        }
        options = [current_truck]
        options.extend(unassigned_trucks)

        seen = set()
        unique_options = []
        for option in options:
            truck_id = option.get("id")
            if truck_id in seen:
                continue
            seen.add(truck_id)
            unique_options.append(option)
        trucks_for_edit[assignment.get("id")] = unique_options

    return render_template(
        "rfid_mappings.html",
        assignments=assignments,
        unassigned_tags=unassigned_tags,
        unassigned_trucks=unassigned_trucks,
        trucks_for_edit=trucks_for_edit,
    )


@app.post("/rfid/tags/import")
def rfid_import_tags():
    db = get_db()
    uploaded = request.files.get("rfid_tags_file")
    if uploaded is None or not (uploaded.filename or "").strip():
        flash("Please choose a CSV or Excel file to upload.", "error")
        return redirect(url_for("rfid_mappings"))

    try:
        raw_rows = parse_materials_upload_rows(uploaded, uploaded.filename)
    except Exception as exc:
        flash(f"Could not read upload: {exc}", "error")
        return redirect(url_for("rfid_mappings"))

    if not raw_rows:
        flash("Uploaded file contains no data rows.", "error")
        return redirect(url_for("rfid_mappings"))

    epc_col, serial_col = resolve_rfid_upload_columns(raw_rows[0][1])
    if not epc_col or not serial_col:
        flash("Missing required columns. Need EPC and serial_number (or serial/sticker_number).", "error")
        return redirect(url_for("rfid_mappings"))

    created = 0
    updated = 0
    skipped = 0
    conflicts = []

    for row_number, row in raw_rows:
        epc = normalize_epc(row.get(epc_col))
        serial_number = normalize_serial_number(row.get(serial_col))
        if not epc or not serial_number:
            skipped += 1
            continue

        try:
            with db.cursor() as cursor:
                cursor.execute("SELECT serial_number FROM rfid_tag_catalog WHERE LOWER(TRIM(epc)) = LOWER(TRIM(%s)) LIMIT 1", (epc,))
                epc_owner = cursor.fetchone()
                if epc_owner and str(epc_owner.get("serial_number") or "").strip().upper() != serial_number:
                    conflicts.append(f"Row {row_number}: EPC {epc} already belongs to serial {epc_owner.get('serial_number')}")
                    continue

                cursor.execute("SELECT epc FROM rfid_tag_catalog WHERE LOWER(TRIM(serial_number)) = LOWER(TRIM(%s)) LIMIT 1", (serial_number,))
                serial_owner = cursor.fetchone()
                existing_epc = normalize_epc(serial_owner.get("epc")) if serial_owner else ""

            upsert_rfid_tag_catalog_row(db, serial_number=serial_number, epc=epc)
            if serial_owner and existing_epc == epc:
                updated += 1
            elif serial_owner:
                updated += 1
            else:
                created += 1
        except Exception as exc:
            conflicts.append(f"Row {row_number}: {exc}")

    if conflicts:
        db.rollback()
        preview = "; ".join(conflicts[:3])
        if len(conflicts) > 3:
            preview += f"; and {len(conflicts) - 3} more"
        flash(f"Import failed due to conflicts. {preview}", "error")
        return redirect(url_for("rfid_mappings"))

    db.commit()
    flash(f"RFID tag import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.", "success")
    return redirect(url_for("rfid_mappings"))


@app.post("/rfid/mappings/assign")
def rfid_assign_mapping():
    db = get_db()
    serial_number = normalize_serial_number(request.form.get("serial_number"))
    truck_id_raw = str(request.form.get("truck_id") or "").strip()
    notes = str(request.form.get("notes") or "").strip()

    if not serial_number or not truck_id_raw.isdigit():
        flash("Serial number and truck are required.", "error")
        return redirect(url_for("rfid_mappings"))

    truck_id = int(truck_id_raw)
    with db.cursor() as cursor:
        cursor.execute("SELECT serial_number FROM rfid_tag_catalog WHERE LOWER(TRIM(serial_number)) = LOWER(TRIM(%s)) LIMIT 1", (serial_number,))
        if not cursor.fetchone():
            flash("Selected serial number does not exist in RFID tag catalog.", "error")
            return redirect(url_for("rfid_mappings"))

        cursor.execute("SELECT id FROM rfid_truck_tag_assignment WHERE LOWER(TRIM(serial_number)) = LOWER(TRIM(%s)) LIMIT 1", (serial_number,))
        if cursor.fetchone():
            flash("This serial number is already assigned to a truck.", "error")
            return redirect(url_for("rfid_mappings"))

        cursor.execute("SELECT id FROM rfid_truck_tag_assignment WHERE truck_id = %s LIMIT 1", (truck_id,))
        if cursor.fetchone():
            flash("This truck is already assigned to another serial number.", "error")
            return redirect(url_for("rfid_mappings"))

        cursor.execute(
            """
            INSERT INTO rfid_truck_tag_assignment (serial_number, truck_id, notes, assigned_at, updated_at)
            VALUES (%s, %s, %s, NOW(), NOW())
            """,
            (serial_number, truck_id, notes),
        )

    sync_epc_map_for_assignment(db, serial_number)
    db.commit()
    flash("RFID serial assigned to truck.", "success")
    return redirect(url_for("rfid_mappings"))


@app.post("/rfid/mappings/<int:assignment_id>/update")
def rfid_update_mapping(assignment_id):
    db = get_db()
    truck_id_raw = str(request.form.get("truck_id") or "").strip()
    notes = str(request.form.get("notes") or "").strip()
    if not truck_id_raw.isdigit():
        flash("Please select a valid truck.", "error")
        return redirect(url_for("rfid_mappings"))
    truck_id = int(truck_id_raw)

    with db.cursor() as cursor:
        cursor.execute("SELECT id, serial_number, truck_id FROM rfid_truck_tag_assignment WHERE id = %s LIMIT 1", (assignment_id,))
        assignment = cursor.fetchone()
        if not assignment:
            flash("Mapping not found.", "error")
            return redirect(url_for("rfid_mappings"))

        current_truck_id = int(assignment.get("truck_id") or 0)
        if truck_id != current_truck_id:
            cursor.execute("SELECT id FROM rfid_truck_tag_assignment WHERE truck_id = %s AND id <> %s LIMIT 1", (truck_id, assignment_id))
            if cursor.fetchone():
                flash("Selected truck is already mapped to another serial.", "error")
                return redirect(url_for("rfid_mappings"))

        cursor.execute(
            """
            UPDATE rfid_truck_tag_assignment
            SET truck_id = %s,
                notes = %s,
                updated_at = NOW()
            WHERE id = %s
            """,
            (truck_id, notes, assignment_id),
        )
        serial_number = str(assignment.get("serial_number") or "").strip()

    sync_epc_map_for_assignment(db, serial_number)
    db.commit()
    flash("RFID mapping updated.", "success")
    return redirect(url_for("rfid_mappings"))


@app.post("/rfid/mappings/<int:assignment_id>/remove")
def rfid_remove_mapping(assignment_id):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT serial_number FROM rfid_truck_tag_assignment WHERE id = %s LIMIT 1", (assignment_id,))
        row = cursor.fetchone()
        if not row:
            flash("Mapping not found.", "error")
            return redirect(url_for("rfid_mappings"))
        serial_number = str(row.get("serial_number") or "").strip()

        cursor.execute("DELETE FROM rfid_truck_tag_assignment WHERE id = %s", (assignment_id,))

    deactivate_epc_map_for_serial(db, serial_number)
    db.commit()
    flash("RFID mapping removed.", "success")
    return redirect(url_for("rfid_mappings"))


@app.get("/rfid/lookup")
def rfid_lookup():
    db = get_db()
    query = str(request.args.get("q") or "").strip()
    rows = search_rfid_assignments(db, query)
    return render_template("rfid_lookup.html", rows=rows, query=query)


@app.post("/tickets/<int:ticket_id>/void")
def void_ticket(ticket_id):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT ticket_number FROM tickets WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()

    if not row:
        flash("Ticket not found.", "error")
        return redirect(request.referrer or url_for("search_tickets"))

    try:
        with db.cursor() as cursor:
            cursor.execute(
                "UPDATE tickets SET active = FALSE WHERE id = %s AND COALESCE(active, TRUE) = TRUE",
                (ticket_id,),
            )
            if cursor.rowcount == 0:
                flash("Ticket is already inactive.", "error")
                db.rollback()
                return redirect(request.referrer or url_for("search_tickets"))
        db.commit()
        flash(f"Ticket {row['ticket_number']} was voided.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Could not void ticket: {exc}", "error")

    return redirect(request.referrer or url_for("search_tickets"))


@app.route("/admin/trucks", methods=["GET", "POST"])
def admin_trucks():
    db = get_db()
    if request.method == "POST":
        truck_number = request.form.get("truck_number", "").strip()
        description = request.form.get("description", "").strip()
        truck_size = request.form.get("truck_size", "").strip()
        hauled_by = request.form.get("hauled_by", "").strip()
        license_plate = request.form.get("license_plate", "").strip()
        if not truck_number:
            flash("Truck number is required.", "error")
            return redirect(url_for("admin_trucks"))
        try:
            active_value = True if is_active_column_boolean(db, "trucks_main") else 1
            with db.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO trucks_main (truck_number, notes, truck_size, trucking_company, license_plate, active) VALUES (%s, %s, %s, %s, %s, %s)",
                    (truck_number, description, truck_size, hauled_by, license_plate, active_value),
                )
            db.commit()
            flash("Truck added.", "success")
        except IntegrityError:
            db.rollback()
            flash("Truck number already exists.", "error")
        return redirect(url_for("admin_trucks"))

    truck_query = request.args.get("q", "").strip()

    with db.cursor() as cursor:
        if truck_query:
            like_query = f"%{truck_query}%"
            cursor.execute(
                """
                SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, license_plate, active
                FROM trucks_main
                WHERE truck_number ILIKE %s
                   OR COALESCE(truck_size, '') ILIKE %s
                   OR COALESCE(trucking_company, '') ILIKE %s
                   OR COALESCE(license_plate, '') ILIKE %s
                ORDER BY truck_number
                """,
                (like_query, like_query, like_query, like_query),
            )
        else:
            cursor.execute(
                "SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, license_plate, active FROM trucks_main ORDER BY truck_number"
            )
        rows = cursor.fetchall()
    return render_template("admin_trucks.html", trucks=rows, truck_query=truck_query)


@app.post("/admin/trucks/<int:truck_id>/toggle")
def toggle_truck(truck_id):
    db = get_db()
    with db.cursor() as cursor:
        if is_active_column_boolean(db, "trucks_main"):
            cursor.execute(
                "UPDATE trucks_main SET active = NOT COALESCE(active, FALSE) WHERE id = %s",
                (truck_id,),
            )
        else:
            cursor.execute(
                "UPDATE trucks_main SET active = CASE WHEN COALESCE(active, 0) = 1 THEN 0 ELSE 1 END WHERE id = %s",
                (truck_id,),
            )
    db.commit() 
    return redirect(url_for("admin_trucks"))


@app.get("/admin/trucks/edit")
def edit_trucks():
    db = get_db()
    truck_query = request.args.get("q", "").strip()
    edit_id_raw = request.args.get("edit_id", "").strip()

    edit_id = int(edit_id_raw) if edit_id_raw.isdigit() else None

    with db.cursor() as cursor:
        if truck_query:
            like_query = f"%{truck_query}%"
            cursor.execute(
                """
                SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, license_plate, active
                FROM trucks_main
                WHERE truck_number ILIKE %s
                   OR COALESCE(truck_size, '') ILIKE %s
                   OR COALESCE(trucking_company, '') ILIKE %s
                   OR COALESCE(license_plate, '') ILIKE %s
                ORDER BY truck_number
                """,
                (like_query, like_query, like_query, like_query),
            )
        else:
            cursor.execute(
                "SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, license_plate, active FROM trucks_main ORDER BY truck_number"
            )
        rows = cursor.fetchall()

    selected_truck = None
    if edit_id is not None:
        selected_truck = next((t for t in rows if t.get("id") == edit_id), None)
        if not selected_truck:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT id, truck_number, notes AS description, truck_size, trucking_company AS hauled_by, license_plate, active
                    FROM trucks_main
                    WHERE id = %s
                    """,
                    (edit_id,),
                )
                selected_truck = cursor.fetchone()

    return render_template(
        "edit_trucks.html",
        trucks=rows,
        selected_truck=selected_truck,
        truck_query=truck_query,
    )


@app.post("/admin/trucks/<int:truck_id>/edit-save")
def save_truck_edit(truck_id):
    db = get_db()
    truck_query = request.form.get("q", "").strip()

    truck_number = request.form.get("truck_number", "").strip()
    description = request.form.get("description", "").strip()
    truck_size = request.form.get("truck_size", "").strip()
    hauled_by = request.form.get("hauled_by", "").strip()
    license_plate = request.form.get("license_plate", "").strip()
    active_raw = request.form.get("active", "").strip().lower()

    if not truck_number:
        flash("Truck number is required.", "error")
        return redirect(url_for("edit_trucks", q=truck_query, edit_id=truck_id))

    active_is_boolean = is_active_column_boolean(db, "trucks_main")
    if active_is_boolean:
        active_value = active_raw in {"1", "true", "yes", "on"}
    else:
        active_value = 1 if active_raw in {"1", "true", "yes", "on"} else 0

    try:
        with db.cursor() as cursor:
            cursor.execute(
                """
                UPDATE trucks_main
                SET truck_number = %s,
                    notes = %s,
                    truck_size = %s,
                    trucking_company = %s,
                    license_plate = %s,
                    active = %s
                WHERE id = %s
                """,
                (truck_number, description, truck_size, hauled_by, license_plate, active_value, truck_id),
            )
            if cursor.rowcount == 0:
                db.rollback()
                flash("Truck not found.", "error")
                return redirect(url_for("edit_trucks", q=truck_query))

        db.commit()
        flash(f"Truck {truck_number} updated.", "success")
    except IntegrityError:
        db.rollback()
        flash("Truck number already exists.", "error")
        return redirect(url_for("edit_trucks", q=truck_query, edit_id=truck_id))
    except Exception as exc:
        db.rollback()
        flash(f"Could not update truck: {exc}", "error")
        return redirect(url_for("edit_trucks", q=truck_query, edit_id=truck_id))

    return redirect(url_for("edit_trucks", q=truck_query, edit_id=truck_id))


@app.get("/admin/materials")
def admin_materials():
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, cat, material AS material_name, cost_per_cy, direction,
            axle_1, tandem, triaxle, axle_4_5, axle_6, semi, hydvac, dirt_in, active
            FROM material_price
            ORDER BY active DESC, direction, material
            """
        )
        rows = cursor.fetchall()
    return render_template("admin_materials.html", materials=rows)


@app.get("/admin/materials/export.pdf")
def export_materials_pdf():
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, cat, material AS material_name, cost_per_cy, direction,
            axle_1, tandem, triaxle, axle_4_5, axle_6, semi, hydvac, dirt_in, active
            FROM material_price
            ORDER BY active DESC, direction, material
            """
        )
        rows = cursor.fetchall()

    pdf_bytes = materials_report_to_pdf_bytes(rows)
    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    filename = f"materials_export_{stamp}.pdf"

    try:
        upload_download_audit_blob(
            category="materials_pdf",
            filename=filename,
            file_bytes=pdf_bytes,
            mimetype="application/pdf",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload materials PDF download: %s", exc)

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.get("/admin/materials/export.csv")
def export_materials_csv():
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, cat, material AS material_name, cost_per_cy, direction,
            axle_1, tandem, triaxle, axle_4_5, axle_6, semi, hydvac, dirt_in, active
            FROM material_price
            ORDER BY id
            """
        )
        rows = cursor.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id",
        "cat",
        "material_name",
        "cost_per_cy",
        "direction",
        "axle_1",
        "tandem",
        "triaxle",
        "axle_4_5",
        "axle_6",
        "semi",
        "hydvac",
        "dirt_in",
        "active",
    ])

    for row in rows:
        writer.writerow([
            row["id"],
            row["cat"],
            row["material_name"],
            row["cost_per_cy"],
            row["direction"],
            row["axle_1"],
            row["tandem"],
            row["triaxle"],
            row["axle_4_5"],
            row["axle_6"],
            row["semi"],
            row["hydvac"],
            row["dirt_in"],
            1 if row["active"] else 0,
        ])

    csv_bytes = output.getvalue().encode("utf-8")
    stamp = app_now().strftime("%Y%m%d_%H%M%S")
    filename = f"materials_export_{stamp}.csv"

    try:
        upload_download_audit_blob(
            category="materials_csv",
            filename=filename,
            file_bytes=csv_bytes,
            mimetype="text/csv",
        )
    except Exception as exc:
        app.logger.warning("Could not audit-upload materials CSV download: %s", exc)

    return send_file(
        io.BytesIO(csv_bytes),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@app.post("/admin/materials/import")
def import_materials_csv():
    db = get_db()
    admin_password = request.form.get("admin_password", "")
    password_ok, password_message = validate_material_admin_password(admin_password)
    if not password_ok:
        flash(f"{password_message} CSV was not imported.", "error")
        return redirect(url_for("admin_materials"))

    uploaded = request.files.get("materials_file")
    if uploaded is None or not (uploaded.filename or "").strip():
        flash("Please choose a CSV or Excel file to upload.", "error")
        return redirect(url_for("admin_materials"))

    try:
        raw_rows = parse_materials_upload_rows(uploaded, uploaded.filename)
    except Exception as exc:
        flash(f"Could not read upload: {exc}", "error")
        return redirect(url_for("admin_materials"))

    if not raw_rows:
        flash("Uploaded file contains no data rows.", "error")
        return redirect(url_for("admin_materials"))

    required_columns = {
        "id",
        "cat",
        "material_name",
        "cost_per_cy",
        "direction",
        "axle_1",
        "tandem",
        "triaxle",
        "axle_4_5",
        "axle_6",
        "semi",
        "hydvac",
        "dirt_in",
        "active",
    }
    axle_keys = [
    "axle_1",
    "tandem",
    "triaxle",
    "axle_4_5",
    "axle_6",
    "semi",
    "hydvac",
    "dirt_in",
    ]

    sample_columns = {str(k).strip().lower() for k in raw_rows[0][1].keys()}
    missing = sorted(required_columns - sample_columns)
    if missing:
        flash("Missing required columns: " + ", ".join(missing), "error")
        return redirect(url_for("admin_materials"))

    use_boolean_active = is_active_column_boolean(db, "material_price")
    parsed_rows = []

    for row_number, row in raw_rows:
        row_map = {str(k).strip().lower(): row.get(k) for k in row.keys()}
        try:
            raw_id = str(row_map.get("id") or "").strip()
            material_id = int(raw_id) if raw_id else None
            cat = int(str(row_map.get("cat") or "").strip())
            material_name = str(row_map.get("material_name") or "").strip()
            cost_per_cy = float(str(row_map.get("cost_per_cy") or "").strip())
            direction = str(row_map.get("direction") or "").strip().upper()
            if not material_name:
                raise ValueError("material_name is required")
            if direction not in {"IN", "OUT"}:
                raise ValueError("direction must be IN or OUT")

            axle_values = []
            for key in axle_keys:
                val = str(row_map.get(key) or "").strip()
                axle_values.append(None if val == "" else float(val))

            active_value = parse_material_active_value(row_map.get("active"), use_boolean_active)

            parsed_rows.append((
                material_id,
                cat,
                material_name,
                direction,
                cost_per_cy,
                *axle_values,
                active_value,
            ))
        except ValueError as exc:
            flash(f"Row {row_number}: {exc}", "error")
            return redirect(url_for("admin_materials"))

    try:
        updated_count = 0
        inserted_count = 0
        with db.cursor() as cursor:
            for parsed in parsed_rows:
                material_id = parsed[0]
                values = parsed[1:]

                if material_id is not None:
                    cursor.execute(
                        """
                        UPDATE material_price
                        SET cat = %s,
                            material = %s,
                            direction = %s,
                            cost_per_cy = %s,
                            axle_1 = %s,
                            tandem = %s,
                            triaxle = %s,
                            axle_4_5 = %s,
                            axle_6 = %s,
                            semi = %s,
                            hydvac = %s,
                            dirt_in = %s,
                            active = %s
                        WHERE id = %s
                        """,
                        (*values, material_id),
                    )

                    if cursor.rowcount > 0:
                        updated_count += cursor.rowcount
                        continue

                cursor.execute(
                    """
                    INSERT INTO material_price (
                        cat, material, cost_per_cy, direction,
                        axle_1, tandem, triaxle, axle_4_5, axle_6, semi, hydvac, dirt_in,
                        active
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    values,
                )
                inserted_count += 1

        db.commit()
        flash(
            f"Materials import complete. Updated {updated_count} row(s), inserted {inserted_count} row(s).",
            "success",
        )
    except Exception as exc:
        db.rollback()
        flash(f"Could not import materials: {exc}", "error")

    return redirect(url_for("admin_materials"))


@app.post("/admin/materials/<int:material_id>/edit")
def edit_material(material_id):
    flash("Direct material editing is disabled. Use CSV upload to update materials.", "error")
    return redirect(url_for("admin_materials"))


@app.route("/admin/customers", methods=["GET", "POST"])
def admin_customers():
    db = get_db()
    if request.method == "POST":
        customer_name = request.form.get("customer_name", "").strip()
        full_address = request.form.get("full_address", "").strip()
        contact_person = request.form.get("contact_person", "").strip()
        phone_number = request.form.get("phone_number", "").strip()
        notes = request.form.get("notes", "").strip()

        if not customer_name:
            flash("Customer name is required.", "error")
            return redirect(url_for("admin_customers"))

        existing_customer = get_customer_by_name(db, customer_name)
        if existing_customer:
            flash("Customer already exists.", "error")
            return redirect(url_for("admin_customers"))

        try:
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO customers (customer_name, full_address, contact_person, phone_number, notes)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (customer_name, full_address, contact_person, phone_number, notes),
                )
            db.commit()
            flash("Customer added.", "success")
        except Exception as exc:
            db.rollback()
            flash(f"Could not add customer: {exc}", "error")

        return redirect(url_for("admin_customers"))

    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, customer_name, full_address, contact_person, phone_number, notes
            FROM customers
            ORDER BY customer_name
            """
        )
        rows = cursor.fetchall()
    return render_template("admin_customers.html", customers=rows)


@app.post("/admin/materials/<int:material_id>/toggle")
def toggle_material(material_id):
    flash("Direct material status changes are disabled. Use CSV upload to update materials.", "error")
    return redirect(url_for("admin_materials"))


if __name__ == "__main__":
    ensure_app_initialized()
    refresh_jobs_on_startup()
    # host = os.getenv("FLASK_RUN_HOST", "127.0.0.1")
    host = os.getenv("FLASK_RUN_HOST", "0.0.0.0")
    port = int(os.getenv("PORT", os.getenv("FLASK_RUN_PORT", "5000")))
    debug = os.getenv("FLASK_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    app.run(debug=debug, host=host, port=port)
    